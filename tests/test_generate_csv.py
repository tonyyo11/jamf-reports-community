"""Integration tests for CSV-backed workbook generation."""

from __future__ import annotations

import csv
import shutil
from pathlib import Path

import openpyxl
import pandas as pd
import pytest


@pytest.mark.integration
def test_generate_dummy_computer_csv_workbook(
    config_factory,
    fixtures_root: Path,
    tmp_path: Path,
    jrc,
) -> None:
    config = config_factory("dummy.yaml")
    config._data["jamf_cli"]["enabled"] = False
    csv_path = fixtures_root / "csv" / "dummy_all_macs.csv"
    out_path = tmp_path / "dummy-report.xlsx"
    historical_dir = tmp_path / "snapshots" / "computers"
    shutil.copytree(fixtures_root / "snapshots" / "computers", historical_dir)

    report_path = jrc.cmd_generate(
        config,
        str(csv_path),
        str(out_path),
        str(historical_dir),
    )

    assert report_path.exists()
    workbook = openpyxl.load_workbook(report_path, data_only=False)
    assert "Device Inventory" in workbook.sheetnames
    assert "Stale Devices" in workbook.sheetnames
    assert "Security Controls" in workbook.sheetnames
    assert "Report Sources" in workbook.sheetnames

    with open(csv_path, encoding="utf-8-sig", newline="") as fh:
        first_row = next(csv.DictReader(fh))
    stale_devices = workbook["Stale Devices"]
    assert stale_devices["A4"].value == "Computer Name"
    assert stale_devices["A5"].value == first_row["Computer Name"]
    if jrc._load_matplotlib():
        assert "Charts" in workbook.sheetnames


@pytest.mark.integration
def test_generate_mobile_csv_workbook(
    config_factory,
    fixtures_root: Path,
    tmp_path: Path,
    jrc,
) -> None:
    config = config_factory("harbor-mobile.yaml")
    config._data["charts"]["enabled"] = True
    config._data["charts"]["embed_in_xlsx"] = True
    config._data["charts"]["os_adoption"]["enabled"] = True
    csv_path = fixtures_root / "csv" / "harbor_mobile_insights_all_devices.csv"
    out_path = tmp_path / "harbor-mobile-report.xlsx"

    report_path = jrc.cmd_generate(config, str(csv_path), str(out_path))

    assert report_path.exists()
    workbook = openpyxl.load_workbook(report_path, data_only=False)
    assert "Mobile Device Inventory" in workbook.sheetnames
    assert "Mobile Stale Devices" in workbook.sheetnames
    assert "Report Sources" in workbook.sheetnames

    inventory_sheet = workbook["Mobile Device Inventory"]
    assert inventory_sheet["A4"].value == "Device Name"
    stale_sheet = workbook["Mobile Stale Devices"]
    assert stale_sheet["A4"].value == "Device Name"
    if jrc._load_matplotlib():
        assert "Charts" in workbook.sheetnames


@pytest.mark.integration
def test_generate_csv_workbook_includes_fleet_drift(
    config_factory,
    fixtures_root: Path,
    tmp_path: Path,
    jrc,
) -> None:
    """A prior CSV snapshot enables the Fleet Drift sheet."""
    config = config_factory("dummy.yaml")
    config._data["jamf_cli"]["enabled"] = False
    config._data["charts"]["enabled"] = False
    csv_path = fixtures_root / "csv" / "dummy_all_macs.csv"
    historical_dir = tmp_path / "snapshots"
    historical_dir.mkdir()

    prior = pd.read_csv(csv_path, dtype=str, encoding="utf-8-sig").fillna("")
    prior.loc[0, "Operating System"] = "14.0"
    prior.to_csv(historical_dir / "dummy_all_macs_2026-04-01.csv", index=False)

    report_path = jrc.cmd_generate(
        config,
        str(csv_path),
        str(tmp_path / "fleet-drift.xlsx"),
        str(historical_dir),
    )

    workbook = openpyxl.load_workbook(report_path, data_only=False)
    assert "Fleet Drift" in workbook.sheetnames
    drift_values = [
        cell.value
        for row in workbook["Fleet Drift"].iter_rows()
        for cell in row
        if cell.value
    ]
    assert "OS Changed" in drift_values


@pytest.mark.integration
def test_generate_csv_workbook_respects_sheets_skip(
    config_factory,
    fixtures_root: Path,
    tmp_path: Path,
    jrc,
) -> None:
    config = config_factory("dummy.yaml")
    config._data["jamf_cli"]["enabled"] = False
    config._data["sheets"] = {
        "skip": ["Device Inventory", "Report Sources", "Charts"],
    }
    csv_path = fixtures_root / "csv" / "dummy_all_macs.csv"

    report_path = jrc.cmd_generate(config, str(csv_path), str(tmp_path / "filtered-report.xlsx"))

    workbook = openpyxl.load_workbook(report_path, data_only=False)
    assert "Device Inventory" not in workbook.sheetnames
    assert "Stale Devices" in workbook.sheetnames
    assert "Report Sources" not in workbook.sheetnames
    assert "Charts" not in workbook.sheetnames


@pytest.mark.integration
def test_generate_csv_workbook_respects_sheets_only_and_precedence(
    config_factory,
    fixtures_root: Path,
    tmp_path: Path,
    jrc,
) -> None:
    config = config_factory("dummy.yaml")
    config._data["jamf_cli"]["enabled"] = False
    config._data["sheets"] = {
        "only": ["Stale Devices", "Report Sources"],
        "skip": ["Stale Devices"],
    }
    csv_path = fixtures_root / "csv" / "dummy_all_macs.csv"

    report_path = jrc.cmd_generate(config, str(csv_path), str(tmp_path / "allowlist-report.xlsx"))

    workbook = openpyxl.load_workbook(report_path, data_only=False)
    assert "Stale Devices" in workbook.sheetnames
    assert "Report Sources" in workbook.sheetnames
    assert "Device Inventory" not in workbook.sheetnames
    assert "Charts" not in workbook.sheetnames
