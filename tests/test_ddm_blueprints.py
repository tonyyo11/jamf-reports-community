"""Tests for v2.1.0 PR 9 — DDM Blueprint gate semantics.

Mirrors ``test_compliance_benchmarks.py`` for the Platform Blueprints and
Platform DDM Status sheets. Confirms the dual-gate (``platform.enabled`` +
``experimental.platform_features_enabled`` + ``has_platform_auth``) blocks
each writer when any leg is open, and confirms a fully-open gate renders
the sheet without raising.

Pre-existing sheet-content tests live in ``test_platform_reports.py``;
this file pins the gating combinations so a future regression that
flipped only one switch would surface here.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock

import openpyxl
import pytest
import xlsxwriter


FIXTURES_ROOT = Path(__file__).resolve().parent / "fixtures"
BP_DIR = FIXTURES_ROOT / "jamf-cli-data" / "blueprint-status"
DDM_DIR = FIXTURES_ROOT / "jamf-cli-data" / "ddm-status"


def _load(path: Path) -> Any:
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def _make_dashboard(
    jrc,
    tmp_path: Path,
    bridge_data: dict,
    *,
    platform_enabled: bool,
    experimental_enabled: bool,
    has_platform_auth: bool,
    is_available: bool = True,
):
    """Build a CoreDashboard with explicit dual-gate state."""
    config = jrc.Config(jrc.Config._WORKSPACE_INIT_DEFAULTS_NAME)
    config._data["platform"]["enabled"] = platform_enabled
    config._data.setdefault("experimental", {})["platform_features_enabled"] = (
        experimental_enabled
    )

    bridge = MagicMock()
    bridge.is_available.return_value = is_available
    bridge.has_platform_auth.return_value = has_platform_auth
    for attr, value in bridge_data.items():
        getattr(bridge, attr).return_value = value

    wb_path = str(tmp_path / "test.xlsx")
    wb = xlsxwriter.Workbook(wb_path, {"remove_timezone": True})
    fmts = jrc._build_formats(wb)
    dashboard = jrc.CoreDashboard(config, bridge, wb, fmts)
    return dashboard, wb, wb_path


# ---------------------------------------------------------------------------
# Platform Blueprints — gate combinations
# ---------------------------------------------------------------------------


def test_blueprints_blocked_when_experimental_off(jrc, tmp_path) -> None:
    """platform.enabled=true + experimental off → skip with gate message."""
    rows = _load(BP_DIR / "platform_blueprint_status_happy.json")
    dashboard, wb, _ = _make_dashboard(
        jrc,
        tmp_path,
        {"blueprint_status": rows},
        platform_enabled=True,
        experimental_enabled=False,
        has_platform_auth=True,
    )
    with pytest.raises(RuntimeError, match="Platform API gate closed"):
        dashboard._write_platform_blueprints()
    wb.close()


def test_blueprints_blocked_when_platform_enabled_off(jrc, tmp_path) -> None:
    """experimental on + platform.enabled off → skip."""
    rows = _load(BP_DIR / "platform_blueprint_status_happy.json")
    dashboard, wb, _ = _make_dashboard(
        jrc,
        tmp_path,
        {"blueprint_status": rows},
        platform_enabled=False,
        experimental_enabled=True,
        has_platform_auth=True,
    )
    with pytest.raises(RuntimeError, match="Platform API gate closed"):
        dashboard._write_platform_blueprints()
    wb.close()


def test_blueprints_blocked_when_profile_not_platform_auth(jrc, tmp_path) -> None:
    """Both flags on, but profile uses oauth2 → skip."""
    rows = _load(BP_DIR / "platform_blueprint_status_happy.json")
    dashboard, wb, _ = _make_dashboard(
        jrc,
        tmp_path,
        {"blueprint_status": rows},
        platform_enabled=True,
        experimental_enabled=True,
        has_platform_auth=False,
    )
    with pytest.raises(RuntimeError, match="Platform API gate closed"):
        dashboard._write_platform_blueprints()
    wb.close()


def test_blueprints_renders_when_full_gate_open(jrc, tmp_path) -> None:
    """All three gates open → sheet writes without raising."""
    rows = _load(BP_DIR / "platform_blueprint_status_happy.json")
    dashboard, wb, wb_path = _make_dashboard(
        jrc,
        tmp_path,
        {"blueprint_status": rows},
        platform_enabled=True,
        experimental_enabled=True,
        has_platform_auth=True,
    )
    dashboard._write_platform_blueprints()
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path)
    assert "Platform Blueprints" in wb_loaded.sheetnames


# ---------------------------------------------------------------------------
# Platform DDM Status — mirror the same gate set
# ---------------------------------------------------------------------------


def test_ddm_blocked_when_experimental_off(jrc, tmp_path) -> None:
    rows = _load(DDM_DIR / "platform_ddm_status_happy.json")
    dashboard, wb, _ = _make_dashboard(
        jrc,
        tmp_path,
        {"ddm_status": rows},
        platform_enabled=True,
        experimental_enabled=False,
        has_platform_auth=True,
    )
    with pytest.raises(RuntimeError, match="Platform API gate closed"):
        dashboard._write_platform_ddm_status()
    wb.close()


def test_ddm_blocked_when_platform_enabled_off(jrc, tmp_path) -> None:
    rows = _load(DDM_DIR / "platform_ddm_status_happy.json")
    dashboard, wb, _ = _make_dashboard(
        jrc,
        tmp_path,
        {"ddm_status": rows},
        platform_enabled=False,
        experimental_enabled=True,
        has_platform_auth=True,
    )
    with pytest.raises(RuntimeError, match="Platform API gate closed"):
        dashboard._write_platform_ddm_status()
    wb.close()


def test_ddm_blocked_when_profile_not_platform_auth(jrc, tmp_path) -> None:
    rows = _load(DDM_DIR / "platform_ddm_status_happy.json")
    dashboard, wb, _ = _make_dashboard(
        jrc,
        tmp_path,
        {"ddm_status": rows},
        platform_enabled=True,
        experimental_enabled=True,
        has_platform_auth=False,
    )
    with pytest.raises(RuntimeError, match="Platform API gate closed"):
        dashboard._write_platform_ddm_status()
    wb.close()


def test_ddm_renders_when_full_gate_open(jrc, tmp_path) -> None:
    rows = _load(DDM_DIR / "platform_ddm_status_happy.json")
    dashboard, wb, wb_path = _make_dashboard(
        jrc,
        tmp_path,
        {"ddm_status": rows},
        platform_enabled=True,
        experimental_enabled=True,
        has_platform_auth=True,
    )
    dashboard._write_platform_ddm_status()
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path)
    assert "Platform DDM Status" in wb_loaded.sheetnames
