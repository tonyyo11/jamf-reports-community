"""Tests for ProtectCLIBridge graduation and the W22 Protect Plans sheet."""

from __future__ import annotations

import json
from pathlib import Path

import pytest


PROTECT_FIXTURE_DIR = "jamf-cli-data/protect-plans"


def _load_fixture(fixtures_root: Path, name: str):
    path = fixtures_root / PROTECT_FIXTURE_DIR / name
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


# ---------------------------------------------------------------------------
# ProtectCLIBridge — class shape and helpers
# ---------------------------------------------------------------------------


def test_protect_bridge_is_subclass_of_jamf_cli_bridge(jrc) -> None:
    assert issubclass(jrc.ProtectCLIBridge, jrc.JamfCLIBridge)


def test_protect_bridge_normalize_bare_array(jrc, fixtures_root) -> None:
    raw = _load_fixture(fixtures_root, "plans_happy.json")
    plans, raw_count = jrc.ProtectCLIBridge._normalize_plans_payload(raw)
    assert isinstance(plans, list) and len(plans) == 3
    assert raw_count == 3
    assert plans[0]["name"] == "Production Default"


def test_protect_bridge_normalize_envelope_shape(jrc, fixtures_root) -> None:
    inner = _load_fixture(fixtures_root, "plans_happy.json")
    envelope = {"nodes": inner, "pageInfo": {"hasNextPage": False, "endCursor": None}}
    plans, raw_count = jrc.ProtectCLIBridge._normalize_plans_payload(envelope)
    assert len(plans) == 3
    assert raw_count == 3
    assert plans[1]["uuid"].startswith("22222222")


def test_protect_bridge_normalize_none_returns_empty(jrc) -> None:
    assert jrc.ProtectCLIBridge._normalize_plans_payload(None) == ([], 0)
    assert jrc.ProtectCLIBridge._normalize_plans_payload("not json") == ([], 0)


def test_classify_protect_error_auth(jrc) -> None:
    cls = jrc.ProtectCLIBridge._classify_protect_error
    assert cls("HTTP 401 unauthorized: client_id missing") == "auth_not_configured"
    assert cls("403 Forbidden — Protect not configured") == "auth_not_configured"


def test_classify_protect_error_network(jrc) -> None:
    cls = jrc.ProtectCLIBridge._classify_protect_error
    assert cls("connection refused while reaching protect API") == "network_error"
    assert cls("ssl: certificate verify failed") == "network_error"


def test_classify_protect_error_unknown_command(jrc) -> None:
    cls = jrc.ProtectCLIBridge._classify_protect_error
    assert cls("unknown command 'plans' for protect") == "unknown_command"


def test_classify_protect_error_general_fallthrough(jrc) -> None:
    cls = jrc.ProtectCLIBridge._classify_protect_error
    assert cls("something happened that we don't classify") == "general"


def test_is_protect_available_returns_bool(jrc, tmp_path) -> None:
    bridge = jrc.ProtectCLIBridge(
        save_output=False, data_dir=str(tmp_path), profile="test"
    )
    # Probe is best-effort; just verify it returns a bool without raising.
    assert isinstance(bridge.is_protect_available(), bool)


# ---------------------------------------------------------------------------
# JamfCLIBridge protect_* shims
# ---------------------------------------------------------------------------


def test_jamf_cli_bridge_protect_methods_are_shims(jrc, tmp_path) -> None:
    """The legacy methods should construct a ProtectCLIBridge under the hood."""
    bridge = jrc.JamfCLIBridge(
        save_output=False, data_dir=str(tmp_path), profile="dummy"
    )
    shim = bridge._protect_shim_bridge()
    assert isinstance(shim, jrc.ProtectCLIBridge)
    # Shim should be cached on subsequent calls.
    assert bridge._protect_shim_bridge() is shim


# ---------------------------------------------------------------------------
# CoreDashboard._write_protect_plans
# ---------------------------------------------------------------------------


class _StubProtectBridge:
    """Minimal bridge stub that returns canned plans_list payloads."""

    def __init__(self, payload, raise_on_call: Exception | None = None) -> None:
        self._payload = payload
        self._raise = raise_on_call

    def plans_list(self):
        if self._raise is not None:
            raise self._raise
        return self._payload

    # Mirror the broader bridge surface used by other CoreDashboard methods, so
    # that incidental attribute access doesn't crash if the dashboard is asked
    # to do more than write Protect Plans. The W22 sheet only needs plans_list.
    def overview(self):
        return []


def _make_core_dashboard(jrc, tmp_path, bridge, *, protect_enabled=True,
                        plans_enabled=True):
    import xlsxwriter

    config = jrc.Config("__workspace_init_defaults__.yaml")
    config._data["protect"]["enabled"] = protect_enabled
    config._data["protect"]["plans"]["enabled"] = plans_enabled

    out_path = tmp_path / "protect.xlsx"
    workbook = xlsxwriter.Workbook(str(out_path), {"remove_timezone": True})
    fmts = jrc._build_formats(workbook)
    dashboard = jrc.CoreDashboard(config, bridge, workbook, fmts)
    return dashboard, workbook, out_path


def test_write_protect_plans_happy_path(jrc, fixtures_root, tmp_path) -> None:
    payload = _load_fixture(fixtures_root, "plans_happy.json")
    bridge = _StubProtectBridge(payload)
    dashboard, workbook, out_path = _make_core_dashboard(jrc, tmp_path, bridge)
    dashboard._write_protect_plans()
    workbook.close()

    import openpyxl
    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb["Protect Plans"]
    # Find the header row (after title rows + status banner).
    header_row = None
    for r in range(1, 15):
        if sheet.cell(row=r, column=1).value == "Plan Name":
            header_row = r
            break
    assert header_row is not None, "Expected 'Plan Name' header row"
    # Three plans, sorted alphabetically.
    names = [
        sheet.cell(row=header_row + i, column=1).value for i in range(1, 4)
    ]
    assert names == ["Engineering Lab", "Executive Strict", "Production Default"]


def test_write_protect_plans_empty_writes_note(jrc, fixtures_root, tmp_path) -> None:
    payload = _load_fixture(fixtures_root, "plans_empty.json")
    bridge = _StubProtectBridge(payload)
    dashboard, workbook, out_path = _make_core_dashboard(jrc, tmp_path, bridge)
    dashboard._write_protect_plans()
    workbook.close()

    from openpyxl import load_workbook
    sheet = load_workbook(out_path)["Protect Plans"]
    cells = [
        sheet.cell(row=r, column=c).value
        for r in range(1, sheet.max_row + 1)
        for c in range(1, sheet.max_column + 1)
    ]
    assert any(v == "No Protect plans configured." for v in cells)


def test_write_protect_plans_disabled_raises(jrc, fixtures_root, tmp_path) -> None:
    payload = _load_fixture(fixtures_root, "plans_happy.json")
    bridge = _StubProtectBridge(payload)
    dashboard, workbook, _ = _make_core_dashboard(
        jrc, tmp_path, bridge, plans_enabled=False
    )
    with pytest.raises(RuntimeError, match="disabled in config"):
        dashboard._write_protect_plans()
    workbook.close()


def test_write_protect_plans_preserves_pascalcase(jrc, fixtures_root, tmp_path) -> None:
    payload = _load_fixture(fixtures_root, "plans_mixed_casing.json")
    bridge = _StubProtectBridge(payload)
    dashboard, workbook, out_path = _make_core_dashboard(jrc, tmp_path, bridge)
    dashboard._write_protect_plans()
    workbook.close()

    import openpyxl
    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb["Protect Plans"]
    header_row = None
    for r in range(1, 15):
        if sheet.cell(row=r, column=1).value == "Plan Name":
            header_row = r
            break
    cec_value = sheet.cell(row=header_row + 1, column=10).value
    # All five PascalCase keys must be preserved verbatim.
    for key in ("MalwareRiskware", "AdversaryTactics", "SystemTampering",
                "FilelessThreats", "Experimental"):
        assert key in cec_value, f"Expected PascalCase key {key} in {cec_value!r}"


def test_write_protect_plans_minimal_does_not_crash(jrc, fixtures_root, tmp_path) -> None:
    payload = _load_fixture(fixtures_root, "plans_minimal.json")
    bridge = _StubProtectBridge(payload)
    dashboard, workbook, out_path = _make_core_dashboard(jrc, tmp_path, bridge)
    dashboard._write_protect_plans()
    workbook.close()

    import openpyxl
    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb["Protect Plans"]
    header_row = None
    for r in range(1, 15):
        if sheet.cell(row=r, column=1).value == "Plan Name":
            header_row = r
            break
    assert sheet.cell(row=header_row + 1, column=1).value == "Minimal Plan"
    # Description / Auto Update fields blank, not crashed.
    assert sheet.cell(row=header_row + 1, column=3).value in ("", None)
