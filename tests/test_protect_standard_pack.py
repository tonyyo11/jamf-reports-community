"""Tests for the W23 Protect standard pack (Computers, Alerts, Insights)."""

from __future__ import annotations

import json
from pathlib import Path

import pytest


def _load(fixtures_root: Path, family: str, name: str):
    with open(fixtures_root / "jamf-cli-data" / family / name, "r", encoding="utf-8") as fh:
        return json.load(fh)


class _StubProtectBridge:
    """Stub that returns canned payloads for the three new resource calls."""

    def __init__(self, computers=None, alerts=None, insights=None):
        self._computers = computers
        self._alerts = alerts
        self._insights = insights

    def computers_list(self):
        return self._computers

    def alerts_list(self):
        return self._alerts

    def insights_list(self):
        return self._insights

    # Existing W22 surface accessed by other dashboard methods. Not exercised
    # here but required so the dashboard can construct without AttributeError.
    def overview(self):
        return []

    def plans_list(self):
        return []

    def analytics_list(self):
        return []


def _make_dashboard(jrc, tmp_path, bridge, *, gates: dict[str, bool] | None = None):
    """Build a CoreDashboard wired with the stub Protect bridge.

    `gates` toggles per-sheet enabled flags (computers/alerts/insights).
    Defaults: protect.enabled=True, all three sheet gates True.
    """
    import xlsxwriter

    flags = {"computers": True, "alerts": True, "insights": True}
    if gates:
        flags.update(gates)

    config = jrc.Config("__workspace_init_defaults__.yaml")
    config._data["protect"]["enabled"] = True
    config._data["protect"]["computers"]["enabled"] = flags["computers"]
    config._data["protect"]["alerts"]["enabled"] = flags["alerts"]
    config._data["protect"]["insights"]["enabled"] = flags["insights"]

    placeholder = jrc.JamfCLIBridge(save_output=False, data_dir=str(tmp_path), profile="")
    out_path = tmp_path / "protect.xlsx"
    workbook = xlsxwriter.Workbook(str(out_path), {"remove_timezone": True})
    fmts = jrc._build_formats(workbook)
    dashboard = jrc.CoreDashboard(
        config, placeholder, workbook, fmts, protect_bridge=bridge
    )
    return dashboard, workbook, out_path


# ---------------------------------------------------------------------------
# Bridge — alerts_list / insights_list method existence
# ---------------------------------------------------------------------------


def test_protect_bridge_exposes_alerts_list_method(jrc, tmp_path) -> None:
    bridge = jrc.ProtectCLIBridge(
        save_output=False, data_dir=str(tmp_path), profile=""
    )
    assert hasattr(bridge, "alerts_list")
    assert callable(bridge.alerts_list)


def test_protect_bridge_exposes_insights_list_method(jrc, tmp_path) -> None:
    bridge = jrc.ProtectCLIBridge(
        save_output=False, data_dir=str(tmp_path), profile=""
    )
    assert hasattr(bridge, "insights_list")
    assert callable(bridge.insights_list)


# ---------------------------------------------------------------------------
# Normalizer — _normalize_protect_payload (bare/envelope/None)
# ---------------------------------------------------------------------------


def test_normalize_protect_payload_bare_array(jrc) -> None:
    rows, total = jrc.CoreDashboard._normalize_protect_payload(
        [{"a": 1}, {"a": 2}], hard_cap=10
    )
    assert rows == [{"a": 1}, {"a": 2}] and total == 2


def test_normalize_protect_payload_envelope_shape(jrc) -> None:
    rows, total = jrc.CoreDashboard._normalize_protect_payload(
        {"nodes": [{"x": 1}], "pageInfo": {"next": None}}, hard_cap=10
    )
    assert rows == [{"x": 1}] and total == 1


def test_normalize_protect_payload_applies_hard_cap(jrc) -> None:
    raw = [{"i": i} for i in range(20)]
    rows, total = jrc.CoreDashboard._normalize_protect_payload(raw, hard_cap=5)
    assert len(rows) == 5
    assert total == 20  # raw count preserves the pre-cap total


def test_normalize_protect_payload_none_and_garbage(jrc) -> None:
    assert jrc.CoreDashboard._normalize_protect_payload(None, hard_cap=5) == ([], 0)
    assert jrc.CoreDashboard._normalize_protect_payload("not-json", hard_cap=5) == ([], 0)


# ---------------------------------------------------------------------------
# Computers sheet
# ---------------------------------------------------------------------------


def test_write_protect_computers_happy(jrc, fixtures_root, tmp_path) -> None:
    payload = _load(fixtures_root, "protect-computers", "computers_happy.json")
    bridge = _StubProtectBridge(computers=payload)
    dashboard, workbook, out_path = _make_dashboard(jrc, tmp_path, bridge)
    dashboard._write_protect_computers()
    workbook.close()

    import openpyxl
    sheet = openpyxl.load_workbook(out_path)["Protect Computers"]
    header_row = None
    for r in range(1, 15):
        if sheet.cell(row=r, column=1).value == "Hostname":
            header_row = r
            break
    assert header_row is not None
    # Sorted alphabetical by Hostname (case-insensitive).
    names = [sheet.cell(row=header_row + i, column=1).value for i in range(1, 4)]
    assert names == ["exec-mbp-09", "lab-mac-01", "prod-server-02"]
    # Plan column populated from nested plan.name.
    plan_col = [sheet.cell(row=header_row + i, column=6).value for i in range(1, 4)]
    assert "Engineering Lab" in plan_col


def test_write_protect_computers_envelope(jrc, fixtures_root, tmp_path) -> None:
    payload = _load(fixtures_root, "protect-computers", "computers_envelope.json")
    bridge = _StubProtectBridge(computers=payload)
    dashboard, workbook, out_path = _make_dashboard(jrc, tmp_path, bridge)
    dashboard._write_protect_computers()
    workbook.close()

    import openpyxl
    sheet = openpyxl.load_workbook(out_path)["Protect Computers"]
    # Find the data row by looking for the envelope hostname.
    found = False
    for r in range(1, 25):
        if sheet.cell(row=r, column=1).value == "envelope-test-01":
            found = True
            break
    assert found, "Envelope payload should yield exactly one Computers row"


def test_write_protect_computers_empty(jrc, fixtures_root, tmp_path) -> None:
    payload = _load(fixtures_root, "protect-computers", "computers_empty.json")
    bridge = _StubProtectBridge(computers=payload)
    dashboard, workbook, out_path = _make_dashboard(jrc, tmp_path, bridge)
    dashboard._write_protect_computers()
    workbook.close()

    import openpyxl
    cells = [
        c.value for r in openpyxl.load_workbook(out_path)["Protect Computers"].iter_rows()
        for c in r
    ]
    assert any(v == "No Protect computers reported." for v in cells)


def test_write_protect_computers_disabled_raises(jrc, fixtures_root, tmp_path) -> None:
    payload = _load(fixtures_root, "protect-computers", "computers_happy.json")
    bridge = _StubProtectBridge(computers=payload)
    dashboard, workbook, _ = _make_dashboard(
        jrc, tmp_path, bridge, gates={"computers": False}
    )
    with pytest.raises(RuntimeError, match="disabled in config"):
        dashboard._write_protect_computers()
    workbook.close()


# ---------------------------------------------------------------------------
# Alerts sheet
# ---------------------------------------------------------------------------


def test_write_protect_alerts_happy(jrc, fixtures_root, tmp_path) -> None:
    payload = _load(fixtures_root, "protect-alerts", "alerts_happy.json")
    bridge = _StubProtectBridge(alerts=payload)
    dashboard, workbook, out_path = _make_dashboard(jrc, tmp_path, bridge)
    dashboard._write_protect_alerts()
    workbook.close()

    import openpyxl
    sheet = openpyxl.load_workbook(out_path)["Protect Alerts"]
    header_row = None
    for r in range(1, 15):
        if sheet.cell(row=r, column=1).value == "Created":
            header_row = r
            break
    assert header_row is not None
    # Three alerts written.
    severities = [
        sheet.cell(row=header_row + i, column=2).value for i in range(1, 4)
    ]
    assert set(severities) == {"high", "medium", "low"}
    # Analytics column joins multiple analytic names.
    analytics_col = [
        sheet.cell(row=header_row + i, column=8).value for i in range(1, 4)
    ]
    assert any(v and "Suspicious Process" in v and "LOLBin" in v for v in analytics_col)


def test_write_protect_alerts_empty(jrc, fixtures_root, tmp_path) -> None:
    payload = _load(fixtures_root, "protect-alerts", "alerts_empty.json")
    bridge = _StubProtectBridge(alerts=payload)
    dashboard, workbook, out_path = _make_dashboard(jrc, tmp_path, bridge)
    dashboard._write_protect_alerts()
    workbook.close()

    import openpyxl
    cells = [
        c.value for r in openpyxl.load_workbook(out_path)["Protect Alerts"].iter_rows()
        for c in r
    ]
    assert any(v == "No Protect alerts reported." for v in cells)


def test_write_protect_alerts_disabled_raises(jrc, fixtures_root, tmp_path) -> None:
    bridge = _StubProtectBridge(alerts=[])
    dashboard, workbook, _ = _make_dashboard(
        jrc, tmp_path, bridge, gates={"alerts": False}
    )
    with pytest.raises(RuntimeError, match="disabled in config"):
        dashboard._write_protect_alerts()
    workbook.close()


# ---------------------------------------------------------------------------
# Insights sheet
# ---------------------------------------------------------------------------


def test_write_protect_insights_happy(jrc, fixtures_root, tmp_path) -> None:
    payload = _load(fixtures_root, "protect-insights", "insights_happy.json")
    bridge = _StubProtectBridge(insights=payload)
    dashboard, workbook, out_path = _make_dashboard(jrc, tmp_path, bridge)
    dashboard._write_protect_insights()
    workbook.close()

    import openpyxl
    sheet = openpyxl.load_workbook(out_path)["Protect Insights"]
    header_row = None
    for r in range(1, 15):
        if sheet.cell(row=r, column=1).value == "Label":
            header_row = r
            break
    assert header_row is not None
    labels = [
        sheet.cell(row=header_row + i, column=1).value for i in range(1, 4)
    ]
    # Sorted alphabetical by Label.
    assert labels == sorted(labels, key=lambda s: s.lower())
    # CIS IDs flatten correctly.
    cis_col = [
        sheet.cell(row=header_row + i, column=9).value for i in range(1, 4)
    ]
    assert any(v and "1.1.1" in v for v in cis_col)


def test_write_protect_insights_empty(jrc, fixtures_root, tmp_path) -> None:
    payload = _load(fixtures_root, "protect-insights", "insights_empty.json")
    bridge = _StubProtectBridge(insights=payload)
    dashboard, workbook, out_path = _make_dashboard(jrc, tmp_path, bridge)
    dashboard._write_protect_insights()
    workbook.close()

    import openpyxl
    cells = [
        c.value for r in openpyxl.load_workbook(out_path)["Protect Insights"].iter_rows()
        for c in r
    ]
    assert any(v == "No Protect insights reported." for v in cells)


def test_write_protect_insights_disabled_raises(jrc, fixtures_root, tmp_path) -> None:
    bridge = _StubProtectBridge(insights=[])
    dashboard, workbook, _ = _make_dashboard(
        jrc, tmp_path, bridge, gates={"insights": False}
    )
    with pytest.raises(RuntimeError, match="disabled in config"):
        dashboard._write_protect_insights()
    workbook.close()
