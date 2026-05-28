"""Tests for `_protect_gate` and the v2.1.0 Protect Threat Overview sheet."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pytest


class _FakeConfig:
    """Minimal Config-shaped double for gating-logic tests."""

    def __init__(self, *, protect_enabled: bool) -> None:
        self._data = {
            "experimental": {"protect_features_enabled": protect_enabled},
        }

    def get(self, *keys: str, default: Any = None) -> Any:
        node: Any = self._data
        for k in keys:
            if not isinstance(node, dict):
                return default
            node = node.get(k, default)
        return node


class _FakeProtectBridge:
    def __init__(self, *, available: bool, protect_available: bool) -> None:
        self._available = available
        self._protect_available = protect_available

    def is_available(self) -> bool:
        return self._available

    def is_protect_available(self) -> bool:
        return self._protect_available


# ---------------------------------------------------------------------------
# _protect_gate()
# ---------------------------------------------------------------------------


def test_protect_gate_false_when_flag_disabled(jrc) -> None:
    config = _FakeConfig(protect_enabled=False)
    bridge = _FakeProtectBridge(available=True, protect_available=True)
    assert jrc._protect_gate(config, bridge) is False


def test_protect_gate_false_when_bridge_is_none(jrc) -> None:
    config = _FakeConfig(protect_enabled=True)
    assert jrc._protect_gate(config, None) is False


def test_protect_gate_false_when_bridge_not_available(jrc) -> None:
    config = _FakeConfig(protect_enabled=True)
    bridge = _FakeProtectBridge(available=False, protect_available=True)
    assert jrc._protect_gate(config, bridge) is False


def test_protect_gate_false_when_protect_not_available(jrc) -> None:
    config = _FakeConfig(protect_enabled=True)
    bridge = _FakeProtectBridge(available=True, protect_available=False)
    assert jrc._protect_gate(config, bridge) is False


def test_protect_gate_true_when_flag_and_bridge_both_ready(jrc) -> None:
    config = _FakeConfig(protect_enabled=True)
    bridge = _FakeProtectBridge(available=True, protect_available=True)
    assert jrc._protect_gate(config, bridge) is True


# ---------------------------------------------------------------------------
# Protect Threat Overview sheet
# ---------------------------------------------------------------------------


class _StubProtectBridge:
    """Stub with the surface CoreDashboard touches for the threat sheet."""

    def __init__(self, *, alerts=None, deep_dive_enabled: bool = True) -> None:
        self._alerts = alerts if alerts is not None else []
        self._deep_dive_enabled = deep_dive_enabled

    def alerts_list(self):
        return self._alerts

    def is_available(self) -> bool:
        return True

    def is_protect_available(self) -> bool:
        return self._deep_dive_enabled

    # Required surface for CoreDashboard to construct without AttributeError.
    def overview(self):
        return []

    def computers_list(self):
        return []

    def plans_list(self):
        return []

    def analytics_list(self):
        return []

    def insights_list(self):
        return []


def _make_dashboard(jrc, tmp_path, bridge, *, experimental_on: bool):
    """Build a CoreDashboard wired with the stub Protect bridge."""
    import xlsxwriter

    config = jrc.Config("__workspace_init_defaults__.yaml")
    config._data["experimental"]["protect_features_enabled"] = experimental_on
    placeholder = jrc.JamfCLIBridge(save_output=False, data_dir=str(tmp_path), profile="")
    out_path = tmp_path / "protect_threat.xlsx"
    workbook = xlsxwriter.Workbook(str(out_path), {"remove_timezone": True})
    fmts = jrc._build_formats(workbook)
    dashboard = jrc.CoreDashboard(
        config, placeholder, workbook, fmts, protect_bridge=bridge
    )
    return dashboard, workbook, out_path


def test_threat_overview_skipped_when_gate_closed(jrc, tmp_path) -> None:
    bridge = _StubProtectBridge(alerts=[{"severity": "high"}])
    dashboard, workbook, _ = _make_dashboard(
        jrc, tmp_path, bridge, experimental_on=False
    )
    with pytest.raises(RuntimeError, match="experimental.protect_features_enabled"):
        dashboard._write_protect_threat_overview()
    workbook.close()


def test_threat_overview_skipped_when_protect_unavailable(jrc, tmp_path) -> None:
    bridge = _StubProtectBridge(alerts=[{"severity": "high"}], deep_dive_enabled=False)
    dashboard, workbook, _ = _make_dashboard(
        jrc, tmp_path, bridge, experimental_on=True
    )
    with pytest.raises(RuntimeError, match="experimental.protect_features_enabled"):
        dashboard._write_protect_threat_overview()
    workbook.close()


def test_threat_overview_renders_when_gate_open(jrc, fixtures_root, tmp_path) -> None:
    payload_path = fixtures_root / "jamf-cli-data" / "protect-alerts" / "alerts_happy.json"
    with open(payload_path, encoding="utf-8") as fh:
        payload = json.load(fh)
    bridge = _StubProtectBridge(alerts=payload)
    dashboard, workbook, out_path = _make_dashboard(
        jrc, tmp_path, bridge, experimental_on=True
    )
    dashboard._write_protect_threat_overview()
    workbook.close()

    import openpyxl
    sheet = openpyxl.load_workbook(out_path)["Protect Threat Overview"]
    header_row = None
    for r in range(1, 15):
        if sheet.cell(row=r, column=1).value == "Device":
            header_row = r
            break
    assert header_row is not None
    severities = [
        str(sheet.cell(row=header_row + i, column=3).value or "").lower()
        for i in range(1, 4)
    ]
    # Severity-sorted: high < medium < low.
    assert severities == ["high", "medium", "low"]


def test_threat_overview_empty_alerts_renders_empty_note(jrc, tmp_path) -> None:
    bridge = _StubProtectBridge(alerts=[])
    dashboard, workbook, out_path = _make_dashboard(
        jrc, tmp_path, bridge, experimental_on=True
    )
    dashboard._write_protect_threat_overview()
    workbook.close()

    import openpyxl
    cells = [
        c.value for r in openpyxl.load_workbook(out_path)["Protect Threat Overview"].iter_rows()
        for c in r
    ]
    assert any(v == "No Protect threat alerts reported." for v in cells)
