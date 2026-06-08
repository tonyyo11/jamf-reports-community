"""Tests for patch-definitions release dates and the Patch Compliance columns."""

from __future__ import annotations

from datetime import date

import openpyxl
import xlsxwriter


def _raise(*args, **kwargs):
    raise RuntimeError("no live jamf-cli")


def test_patch_definitions_parses_committed_fixture(jrc, fixtures_root, monkeypatch) -> None:
    bridge = jrc.JamfCLIBridge(
        save_output=False,
        data_dir=str(fixtures_root / "jamf-cli-data"),
        profile="dummy",
        use_cached_data=True,
    )
    # Force the live call to fail so the committed per-title cache is read.
    monkeypatch.setattr(bridge, "_run", _raise)

    defs = bridge.patch_definitions("2")
    assert isinstance(defs, list)
    assert defs[0]["version"] == "151.0.2"
    assert defs[0]["releaseDate"] == "2026-05-26T13:48:54Z"
    assert bridge.source_info("patch-definitions:2")["mode"] == "cached-fallback"


def test_patch_definitions_blank_id_returns_empty(jrc, fixtures_root) -> None:
    bridge = jrc.JamfCLIBridge(save_output=False, data_dir=str(fixtures_root / "jamf-cli-data"))
    assert bridge.patch_definitions("") == []


def test_latest_definition_exact_version_match(jrc) -> None:
    definitions = [
        {"version": "151.0.2", "releaseDate": "2026-05-26T13:48:54Z", "absoluteOrderId": "0"},
        {"version": "151.0.1", "releaseDate": "2026-05-21T13:32:15Z", "absoluteOrderId": "1"},
    ]
    version, release = jrc.JamfCLIBridge._latest_definition_date(definitions, "151.0.1")
    assert version == "151.0.1"
    assert release == "2026-05-21T13:32:15Z"


def test_latest_definition_falls_back_to_order_zero(jrc) -> None:
    definitions = [
        {"version": "151.0.2", "releaseDate": "2026-05-26T13:48:54Z", "absoluteOrderId": "0"},
        {"version": "151.0.1", "releaseDate": "2026-05-21T13:32:15Z", "absoluteOrderId": "1"},
    ]
    # latest="999.0" matches nothing -> absoluteOrderId "0" entry.
    version, release = jrc.JamfCLIBridge._latest_definition_date(definitions, "999.0")
    assert version == "151.0.2"
    assert release == "2026-05-26T13:48:54Z"


def test_collect_patch_release_dates_merged_shape(jrc, tmp_path, monkeypatch) -> None:
    bridge = jrc.JamfCLIBridge(save_output=True, data_dir=str(tmp_path / "data"))
    monkeypatch.setattr(
        bridge, "patch_status",
        lambda: [{"id": "2", "title": "Mozilla Firefox", "latest": "151.0.2"}],
    )
    monkeypatch.setattr(
        bridge, "patch_definitions",
        lambda tid: [
            {"version": "151.0.2", "releaseDate": "2026-05-26T13:48:54Z", "absoluteOrderId": "0"},
        ],
    )

    merged = bridge.collect_patch_release_dates()
    assert merged == [{
        "title_id": "2",
        "title": "Mozilla Firefox",
        "latest_version": "151.0.2",
        "release_date": "2026-05-26T13:48:54Z",
    }]
    # A snapshot file was written under the dedicated kind dir.
    written = list((tmp_path / "data" / "patch-release-dates").glob("patch-release-dates_*.json"))
    assert len(written) == 1


def test_patch_release_dates_reader_returns_empty_when_absent(jrc, tmp_path) -> None:
    bridge = jrc.JamfCLIBridge(save_output=False, data_dir=str(tmp_path / "data"))
    assert bridge.patch_release_dates() == []


class _PatchBridge:
    """Minimal bridge double for the Patch Compliance sheet."""

    def __init__(self, release_dates):
        self._release_dates = release_dates

    def patch_status(self):
        return [
            {"id": "2", "title": "Mozilla Firefox", "latest": "151.0.2",
             "on_latest": 8, "on_other": 2, "total": 10, "compliance_pct": "80%"},
        ]

    def patch_summaries(self):
        raise RuntimeError("no summaries")

    def patch_release_dates(self):
        return self._release_dates

    def device_compliance(self, stale_days):
        raise RuntimeError("no compliance")


def _patch_sheet(jrc, config_factory, bridge, tmp_path):
    config = config_factory("dummy.yaml")
    out = tmp_path / "patch.xlsx"
    wb = xlsxwriter.Workbook(str(out))
    fmts = jrc._build_formats(wb)
    dash = jrc.CoreDashboard(config, bridge, wb, fmts)
    dash._write_patch()
    wb.close()
    return openpyxl.load_workbook(out)["Patch Compliance"]


def test_patch_sheet_gains_columns_when_snapshot_present(jrc, config_factory, tmp_path) -> None:
    release = [{
        "title_id": "2", "title": "Mozilla Firefox",
        "latest_version": "151.0.2", "release_date": "2026-05-26T13:48:54Z",
    }]
    sheet = _patch_sheet(jrc, config_factory, _PatchBridge(release), tmp_path)
    values = [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value]
    assert "Latest Released" in values
    assert "Days Behind" in values
    # 80% compliance < 100% -> Days Behind populated from 2026-05-26.
    expected_days = max(0, (date.today() - date(2026, 5, 26)).days)
    assert str(expected_days) in values
    assert "2026-05-26" in values


def test_patch_sheet_unchanged_when_snapshot_absent(jrc, config_factory, tmp_path) -> None:
    sheet = _patch_sheet(jrc, config_factory, _PatchBridge([]), tmp_path)
    values = [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value]
    assert "Latest Released" not in values
    assert "Days Behind" not in values
    assert "Software Title" in values


def test_collect_skip_excludes_patch_definitions(jrc, config_factory) -> None:
    config = config_factory("dummy.yaml")
    config._data["jamf_cli"]["collect_skip"] = ["patch_definitions"]
    skip = jrc._normalized_skip_types(config)
    assert "patch-definitions" in skip


def test_collect_skip_drops_patch_release_dates_command(jrc, config_factory) -> None:
    config = config_factory("dummy.yaml")
    config._data["jamf_cli"]["collect_skip"] = ["patch-definitions"]

    class _Stub:
        def __getattr__(self, name):
            return lambda *a, **k: None

    commands = jrc._collect_jamf_cli_commands(config, _Stub(), live_overview_allowed=True)
    labels = [label for label, _ in commands]
    assert "Patch Release Dates" not in labels
