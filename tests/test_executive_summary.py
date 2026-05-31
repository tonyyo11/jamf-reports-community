"""Tests for the v2.1.1 Executive Summary aggregation and rendering.

Covers:
- `_executive_summary_metrics` with full, partial, and empty inputs.
- `_exec_security_score` weighting + renormalization (FV/SIP/Firewall, no GK).
- `_exec_patch_compliance` fleet ratio across both patch-status shapes.
- `CoreDashboard._write_executive_summary` writes "Executive Summary" as the
  first sheet, omits rows whose source is absent, and never raises on partial
  data (unlike the peer dashboards that raise → skip).
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest
import xlsxwriter


_FULL_SECURITY = [
    {
        "section": "summary",
        "data": {
            "total_devices": 100,
            "filevault_encrypted": 90,
            "sip_enabled": 80,
            "firewall_enabled": 70,
            "gatekeeper_enabled": 95,
        },
    },
    {"section": "os_version", "os_version": "15.7.3", "count": 100, "pct": "100%"},
]

_FULL_PATCH = [
    {"title": "Firefox", "on_latest": 80, "on_other": 20, "total": 100},
    {"title": "Chrome", "on_latest": 50, "on_other": 50, "total": 100},
]

_FULL_COMPLIANCE = [
    {"name": "a", "stale": False},
    {"name": "b", "stale": False},
    {"name": "c", "stale": True},
]


# ---------------------------------------------------------------------------
# Pure aggregation — full data
# ---------------------------------------------------------------------------


def test_metrics_full_data(jrc) -> None:
    m = jrc._executive_summary_metrics(_FULL_SECURITY, _FULL_PATCH, _FULL_COMPLIANCE)
    assert m["total_devices"] == 100
    assert m["controls"]["filevault"]["pct"] == pytest.approx(0.90)
    assert m["controls"]["sip"]["pct"] == pytest.approx(0.80)
    assert m["controls"]["firewall"]["pct"] == pytest.approx(0.70)
    assert m["controls"]["gatekeeper"]["pct"] == pytest.approx(0.95)
    # Fleet patch compliance = (80+50)/(100+100) = 0.65
    assert m["patch_compliance"] == pytest.approx(0.65)
    assert m["stale_count"] == 1
    assert m["active_count"] == 2
    # Score = mean of FV/SIP/Firewall (equal 15/15/15 weights), GK excluded.
    assert m["security_score"] == pytest.approx((0.90 + 0.80 + 0.70) / 3)


def test_score_excludes_gatekeeper(jrc) -> None:
    """Gatekeeper is shown but must not enter the weighted score."""
    controls = {
        "filevault": {"pct": 1.0},
        "sip": {"pct": 1.0},
        "firewall": {"pct": 1.0},
        "gatekeeper": {"pct": 0.0},  # would drag score down if included
    }
    assert jrc._exec_security_score(controls) == pytest.approx(1.0)


def test_score_renormalizes_over_present_controls(jrc) -> None:
    """A missing control drops from the denominator; score uses the rest."""
    controls = {
        "filevault": {"pct": 0.80},
        "sip": {"pct": 0.60},
        "firewall": {"pct": None},  # absent — excluded from numerator + denom
    }
    assert jrc._exec_security_score(controls) == pytest.approx((0.80 + 0.60) / 2)


def test_score_none_when_no_weighted_control(jrc) -> None:
    controls = {
        "filevault": {"pct": None},
        "sip": {"pct": None},
        "firewall": {"pct": None},
    }
    assert jrc._exec_security_score(controls) is None


# ---------------------------------------------------------------------------
# Pure aggregation — partial / empty data (graceful degradation)
# ---------------------------------------------------------------------------


def test_metrics_security_only(jrc) -> None:
    """Patch + compliance absent → those metrics are None, security present."""
    m = jrc._executive_summary_metrics(_FULL_SECURITY, None, None)
    assert m["total_devices"] == 100
    assert m["controls"]["filevault"]["pct"] == pytest.approx(0.90)
    assert m["patch_compliance"] is None
    assert m["stale_count"] is None
    assert m["active_count"] is None
    assert m["security_score"] is not None


def test_metrics_compliance_supplies_total_when_security_absent(jrc) -> None:
    """With no security report, device-compliance backfills total + stale."""
    m = jrc._executive_summary_metrics(None, None, _FULL_COMPLIANCE)
    assert m["total_devices"] == 3
    assert m["stale_count"] == 1
    assert m["active_count"] == 2
    assert m["controls"]["filevault"]["pct"] is None
    assert m["security_score"] is None


def test_metrics_all_empty(jrc) -> None:
    m = jrc._executive_summary_metrics(None, None, None)
    assert m["total_devices"] == 0
    assert m["patch_compliance"] is None
    assert m["stale_count"] is None
    assert m["security_score"] is None
    assert all(c["pct"] is None for c in m["controls"].values())


def test_metrics_garbage_inputs_do_not_raise(jrc) -> None:
    m = jrc._executive_summary_metrics("nope", {"x": 1}, 42)
    assert m["total_devices"] == 0
    assert m["security_score"] is None


# ---------------------------------------------------------------------------
# Patch compliance — both shapes + edge cases
# ---------------------------------------------------------------------------


def test_patch_compliance_installed_shape(jrc) -> None:
    """Pre-v1.4 installed/total shape sums correctly."""
    titles = [
        {"title": "A", "installed": 9, "total": 10},
        {"title": "B", "installed": 1, "total": 10},
    ]
    assert jrc._exec_patch_compliance(titles) == pytest.approx(10 / 20)


def test_patch_compliance_derives_total_when_zero(jrc) -> None:
    titles = [{"title": "A", "on_latest": 3, "on_other": 1, "total": 0}]
    assert jrc._exec_patch_compliance(titles) == pytest.approx(3 / 4)


def test_patch_compliance_empty_is_none(jrc) -> None:
    assert jrc._exec_patch_compliance([]) is None
    assert jrc._exec_patch_compliance(None) is None


# ---------------------------------------------------------------------------
# xlsx writer — first sheet, omits absent rows, never raises on partial data
# ---------------------------------------------------------------------------


def _build_dashboard(jrc, tmp_path: Path, security, patch, compliance):
    config = jrc.Config(jrc.Config._WORKSPACE_INIT_DEFAULTS_NAME)
    bridge = MagicMock()
    bridge.security_report.return_value = security
    bridge.patch_status.return_value = patch
    bridge.device_compliance.return_value = compliance
    wb_path = str(tmp_path / "exec.xlsx")
    wb = xlsxwriter.Workbook(wb_path, {"remove_timezone": True})
    fmts = jrc._build_formats(wb)
    dashboard = jrc.CoreDashboard(config, bridge, wb, fmts)
    return dashboard, wb, wb_path


def _metric_map(ws) -> dict:
    out = {}
    for r in range(5, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if label:
            out[label] = ws.cell(row=r, column=2).value
    return out


def test_exec_summary_is_first_sheet(jrc) -> None:
    config = jrc.Config(jrc.Config._WORKSPACE_INIT_DEFAULTS_NAME)
    bridge = MagicMock()
    dashboard = jrc.CoreDashboard(config, bridge, MagicMock(), {})
    plan = dashboard.sheet_plan()
    assert plan[0][0] == "Executive Summary"


def test_write_exec_summary_full(jrc, tmp_path) -> None:
    dashboard, wb, wb_path = _build_dashboard(
        jrc, tmp_path, _FULL_SECURITY, _FULL_PATCH, _FULL_COMPLIANCE
    )
    dashboard._write_executive_summary()
    wb.close()

    book = openpyxl.load_workbook(wb_path, data_only=False)
    assert book.sheetnames[0] == "Executive Summary"
    ws = book["Executive Summary"]
    metrics = _metric_map(ws)
    assert metrics["Total Devices"] == 100
    assert metrics["Stale Devices"] == 1
    assert metrics["Active Devices"] == 2
    # Percentage cells are stored as 0-1 floats.
    assert metrics["FileVault %"] == pytest.approx(0.90)
    assert metrics["Patch Fleet Compliance %"] == pytest.approx(0.65)
    assert "Security Score" in metrics


def test_write_exec_summary_partial_omits_absent_rows(jrc, tmp_path) -> None:
    """Security-only run: patch + stale rows omitted, no exception raised.

    ``None`` (not ``[]``) models an absent source — an empty list means the
    report ran and returned zero rows, which is a real 0, not 'no data'.
    """
    dashboard, wb, wb_path = _build_dashboard(jrc, tmp_path, _FULL_SECURITY, None, None)
    dashboard._write_executive_summary()
    wb.close()

    ws = openpyxl.load_workbook(wb_path)["Executive Summary"]
    metrics = _metric_map(ws)
    assert "FileVault %" in metrics
    assert "Patch Fleet Compliance %" not in metrics
    assert "Stale Devices" not in metrics


def test_write_exec_summary_all_absent_renders_placeholder(jrc, tmp_path) -> None:
    """Every source absent (None) → a 'No jamf-cli data' note, not a raise."""
    dashboard, wb, wb_path = _build_dashboard(jrc, tmp_path, None, None, None)
    dashboard._write_executive_summary()  # must not raise
    wb.close()

    ws = openpyxl.load_workbook(wb_path)["Executive Summary"]
    first_label = ws.cell(row=5, column=1).value
    assert first_label == "No jamf-cli data available"


def test_write_exec_summary_tolerates_fetch_failure(jrc, tmp_path) -> None:
    """A bridge call raising is caught; the sheet still renders what it can."""
    config = jrc.Config(jrc.Config._WORKSPACE_INIT_DEFAULTS_NAME)
    bridge = MagicMock()
    bridge.security_report.return_value = _FULL_SECURITY
    bridge.patch_status.side_effect = RuntimeError("patch report unavailable")
    bridge.device_compliance.return_value = _FULL_COMPLIANCE
    wb_path = str(tmp_path / "exec.xlsx")
    wb = xlsxwriter.Workbook(wb_path, {"remove_timezone": True})
    fmts = jrc._build_formats(wb)
    dashboard = jrc.CoreDashboard(config, bridge, wb, fmts)

    dashboard._write_executive_summary()  # must not raise
    wb.close()

    metrics = _metric_map(openpyxl.load_workbook(wb_path)["Executive Summary"])
    assert metrics["Total Devices"] == 100
    assert "Patch Fleet Compliance %" not in metrics
    assert metrics["Stale Devices"] == 1


# ---------------------------------------------------------------------------
# HTML render — top card row present, score consistent with xlsx
# ---------------------------------------------------------------------------


def test_html_executive_summary_card_row(jrc) -> None:
    report = jrc.HtmlReport.__new__(jrc.HtmlReport)
    html = report._render_executive_summary(_FULL_SECURITY)
    assert "Executive Summary" in html
    assert "Security Score" in html
    assert "Total Devices" in html
    # Score = mean of FV/SIP/Firewall (90/80/70) = 80% (GK excluded).
    assert "80%" in html
    assert "90.0%" in html  # FileVault KPI


def test_html_executive_summary_count_only_shape_matches_xlsx(jrc) -> None:
    """The documented v1.16.1 shape omits *_pct keys — HTML must still score.

    The HTML card and the xlsx sheet share ``_exec_security_metrics``, so a
    counts-only report yields identical FileVault % and Security Score in both,
    rather than the HTML degrading to 0% / N/A.
    """
    counts_only = [
        {
            "section": "summary",
            "data": {
                "total_devices": 100,
                "filevault_encrypted": 90,
                "sip_enabled": 80,
                "firewall_enabled": 70,
                "gatekeeper_enabled": 95,
            },
        }
    ]
    report = jrc.HtmlReport.__new__(jrc.HtmlReport)
    html = report._render_executive_summary(counts_only)
    assert "90.0%" in html  # count-derived FileVault, not 0.0%
    assert "80%" in html     # count-derived Security Score, not N/A

    xlsx_score = jrc._executive_summary_metrics(counts_only, None, None)["security_score"]
    assert xlsx_score == pytest.approx(0.80)  # same value the xlsx sheet writes


def test_html_executive_summary_empty_when_no_data(jrc) -> None:
    report = jrc.HtmlReport.__new__(jrc.HtmlReport)
    assert report._render_executive_summary([]) == ""
    assert report._render_executive_summary(None) == ""
