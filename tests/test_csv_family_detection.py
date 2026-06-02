"""Tests for deterministic CSV family detection, continuation-row dropping,
mobile scaffolding, and the scaffold hint-order tiebreak."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

# config.example.yaml mobile_columns values — these now also appear in modern
# (11.28+) computer exports, which is the family-misdetection trigger (Bug 1).
EXAMPLE_MOBILE_COLUMNS = {
    "device_name": "Display Name",
    "serial_number": "Serial Number",
    "operating_system": "OS Version",
    "last_checkin": "Last Inventory Update",
    "email": "Email Address",
    "model": "Model",
    "device_family": "Device Family",
    "managed": "Managed",
    "supervised": "Supervised",
}


def _headers(path: Path, sep: str = ",") -> list[str]:
    return pd.read_csv(path, nrows=0, sep=sep, encoding="utf-8-sig").columns.tolist()


# ---------------------------------------------------------------------------
# Discriminator-based family detection
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "fixture, expected",
    [
        ("jamf1128_computers_builtin.csv", "computers"),
        ("jamf1128_mobile_builtin.csv", "mobile"),
        ("jamf1128_computers_exportonly.csv", "computers"),
        ("jamf1128_mobile_exportonly.csv", "mobile"),
    ],
)
def test_detect_family_from_builtin_and_exportonly(
    jrc, fixtures_root: Path, fixture: str, expected: str
) -> None:
    headers = _headers(fixtures_root / "csv" / fixture)
    assert jrc._detect_csv_family_from_headers(headers) == expected


def test_no_discriminator_headers_returns_none(jrc) -> None:
    assert jrc._detect_csv_family_from_headers(["foo", "bar"]) is None


def test_school_csv_is_not_misdetected(jrc, fixtures_root: Path) -> None:
    headers = _headers(fixtures_root / "csv" / "harboredu_school_devices.csv", sep=";")
    assert jrc._detect_csv_family_from_headers(headers) is None


def test_populated_mobile_columns_do_not_misdetect_computer_csv(
    jrc, fixtures_root: Path
) -> None:
    """Bug 1: populated mobile_columns must not route a computer CSV to mobile."""
    config = jrc.Config(str(fixtures_root / "config" / "dummy.yaml"))
    config._data["mobile_columns"] = dict(EXAMPLE_MOBILE_COLUMNS)
    headers = _headers(fixtures_root / "csv" / "jamf1128_computers_builtin.csv")
    assert jrc._guess_report_family_from_headers(config, headers) == "computers"


def test_compliance_refinement(jrc, fixtures_root: Path) -> None:
    config = jrc.Config(str(fixtures_root / "config" / "dummy.yaml"))
    headers = _headers(fixtures_root / "csv" / "jamf1128_computers_builtin.csv")
    headers = headers + ["Failed MSCP Results Count"]
    config._data["compliance"]["failures_count_column"] = "Failed MSCP Results Count"
    assert jrc._guess_report_family_from_headers(config, headers) == "compliance"


def test_compliance_refinement_only_when_columns_present(
    jrc, fixtures_root: Path
) -> None:
    config = jrc.Config(str(fixtures_root / "config" / "dummy.yaml"))
    headers = _headers(fixtures_root / "csv" / "jamf1128_computers_builtin.csv")
    config._data["compliance"]["failures_count_column"] = "Not In This CSV"
    assert jrc._guess_report_family_from_headers(config, headers) == "computers"


def test_mobile_guess_unaffected_by_compliance_refinement(
    jrc, fixtures_root: Path
) -> None:
    config = jrc.Config(str(fixtures_root / "config" / "dummy.yaml"))
    headers = _headers(fixtures_root / "csv" / "jamf1128_mobile_builtin.csv")
    config._data["compliance"]["failures_count_column"] = "Failed MSCP Results Count"
    assert jrc._guess_report_family_from_headers(config, headers) == "mobile"


# ---------------------------------------------------------------------------
# Continuation-row dropping (Bug 2)
# ---------------------------------------------------------------------------


def test_drop_continuation_rows_exportonly_computer(jrc, fixtures_root: Path) -> None:
    df = pd.read_csv(
        fixtures_root / "csv" / "jamf1128_computers_exportonly.csv",
        dtype=str,
        encoding="utf-8-sig",
    ).fillna("")
    filtered, dropped = jrc._drop_continuation_rows(df, "Computer Name")
    assert len(filtered) == 4
    assert dropped == 12


def test_drop_continuation_rows_exportonly_mobile(jrc, fixtures_root: Path) -> None:
    df = pd.read_csv(
        fixtures_root / "csv" / "jamf1128_mobile_exportonly.csv",
        dtype=str,
        encoding="utf-8-sig",
    ).fillna("")
    filtered, dropped = jrc._drop_continuation_rows(df, "Display Name")
    assert len(filtered) == 4
    assert dropped == 12


def test_drop_continuation_rows_builtin_drops_nothing(
    jrc, fixtures_root: Path
) -> None:
    df = pd.read_csv(
        fixtures_root / "csv" / "jamf1128_computers_builtin.csv",
        dtype=str,
        encoding="utf-8-sig",
    ).fillna("")
    filtered, dropped = jrc._drop_continuation_rows(df, "Computer Name")
    assert len(filtered) == 4
    assert dropped == 0


def test_drop_continuation_rows_missing_identity_column_no_drop(
    jrc, fixtures_root: Path
) -> None:
    df = pd.read_csv(
        fixtures_root / "csv" / "jamf1128_computers_exportonly.csv",
        dtype=str,
        encoding="utf-8-sig",
    ).fillna("")
    filtered, dropped = jrc._drop_continuation_rows(df, "No Such Column")
    assert len(filtered) == len(df)
    assert dropped == 0


# ---------------------------------------------------------------------------
# Hint-order tiebreak (Bug 4)
# ---------------------------------------------------------------------------


def test_exact_match_respects_hint_order(jrc) -> None:
    """'Last Check-in' (hint index 0) beats 'Last Inventory Update' (index 3)."""
    check_in = jrc._column_match_score("Last Check-in", "last_checkin")
    inventory = jrc._column_match_score("Last Inventory Update", "last_checkin")
    assert check_in > inventory


def test_best_header_match_picks_first_hint(jrc) -> None:
    best, _ = jrc._best_header_match(
        ["Last Inventory Update", "Last Check-in"], "last_checkin"
    )
    assert best == "Last Check-in"


# ---------------------------------------------------------------------------
# Family-aware scaffolding (Bug 3)
# ---------------------------------------------------------------------------


def test_scaffold_computer_csv_populates_columns_blanks_mobile(
    jrc, fixtures_root: Path, tmp_path: Path
) -> None:
    import yaml

    out_path = tmp_path / "computer.yaml"
    jrc.cmd_scaffold(
        str(fixtures_root / "csv" / "jamf1128_computers_builtin.csv"), str(out_path)
    )
    data = yaml.safe_load(out_path.read_text(encoding="utf-8"))
    assert data["columns"]["last_checkin"] == "Last Check-in"
    assert data["columns"]["computer_name"] == "Computer Name"
    assert all(value == "" for value in data["mobile_columns"].values())


def test_scaffold_mobile_csv_populates_mobile_blanks_columns(
    jrc, fixtures_root: Path, tmp_path: Path
) -> None:
    import yaml

    out_path = tmp_path / "mobile.yaml"
    jrc.cmd_scaffold(
        str(fixtures_root / "csv" / "jamf1128_mobile_builtin.csv"), str(out_path)
    )
    data = yaml.safe_load(out_path.read_text(encoding="utf-8"))
    assert data["mobile_columns"]["device_name"] == "Display Name"
    assert data["mobile_columns"]["operating_system"] == "OS Version"
    assert all(value == "" for value in data["columns"].values())


def test_scaffold_mobile_csv_excludes_multivalue_columns(
    jrc, fixtures_root: Path, tmp_path: Path, capsys
) -> None:
    out_path = tmp_path / "mobile.yaml"
    jrc.cmd_scaffold(
        str(fixtures_root / "csv" / "jamf1128_mobile_exportonly.csv"), str(out_path)
    )
    captured = capsys.readouterr().out
    # Multi-value per-item columns must not be suggested as custom EAs.
    for column in (
        "Application Title",
        "Certificate Issuer",
        "Profile Name",
    ):
        assert f"name: '{column}'" not in captured


def test_scaffold_mobile_config_passes_check(
    jrc, fixtures_root: Path, tmp_path: Path
) -> None:
    out_path = tmp_path / "mobile.yaml"
    csv_path = fixtures_root / "csv" / "jamf1128_mobile_builtin.csv"
    jrc.cmd_scaffold(str(csv_path), str(out_path))
    config = jrc.Config(str(out_path))
    config._data["jamf_cli"]["enabled"] = False
    # Should not raise.
    jrc.cmd_check(config, str(csv_path))


def test_multivalue_predicate_matches_enrollment_method_prefix(jrc) -> None:
    assert jrc._is_mobile_multivalue_column("Enrollment Method: PreStage enrollment")
    assert jrc._is_mobile_multivalue_column("Group")
    assert not jrc._is_mobile_multivalue_column("Battery Health")


# ---------------------------------------------------------------------------
# End-to-end generate regression (Bug 1 + Bug 2)
# ---------------------------------------------------------------------------


@pytest.mark.integration
def test_generate_exportonly_computer_routes_to_computer_sheets(
    config_factory, fixtures_root: Path, tmp_path: Path, jrc
) -> None:
    config = config_factory("dummy.yaml")
    config._data["jamf_cli"]["enabled"] = False
    config._data["charts"]["enabled"] = False
    csv_path = fixtures_root / "csv" / "jamf1128_computers_exportonly.csv"
    out_path = tmp_path / "exportonly-computer.xlsx"

    report_path = jrc.cmd_generate(config, str(csv_path), str(out_path))

    workbook = openpyxl.load_workbook(report_path, data_only=False)
    assert "Device Inventory" in workbook.sheetnames
    assert "Security Controls" in workbook.sheetnames
    assert "Mobile Device Inventory" not in workbook.sheetnames
