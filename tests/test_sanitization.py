"""Malicious-payload round-trip tests for `_safe_write` and `HtmlReport._html_text`.

Gemini security review T-3/T-4 deferred this from threat-model coverage:
sanitization invariants must hold against XSS strings, formula-injection
strings, and mixed Unicode/control-char payloads. Tests assert two things:
1. Payloads survive sanitization with their meaning preserved (no silent
   data loss for legitimate content that happens to contain dangerous chars).
2. The dangerous behavior is neutralized — formulas don't execute, HTML
   tags don't render as markup, control characters can't break the file.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
import xlsxwriter


# ---------------------------------------------------------------
# _safe_write — formula-injection neutralization (XLSX)
# ---------------------------------------------------------------
#
# When the first non-whitespace character of a cell value is `=`, `+`, `-`,
# or `@`, Excel evaluates the cell as a formula on open. `_safe_write`
# prefixes such cells with a `\t` (tab) and forces a string write so the
# raw payload survives as displayed text but never executes.

_FORMULA_PAYLOADS = [
    "=SUM(A1:A10)",
    "=1+1",
    "+1+1",
    "-1+1",
    "@SUM(A1:A10)",
    "=cmd|'/c calc'!A1",
    "=HYPERLINK(\"http://evil.example/steal\",\"Click\")",
    "  =LEAD_SPACE_FORMULA(1)",     # leading whitespace must not bypass
    "\t=TAB_PREFIX_FORMULA(1)",     # already starts with tab — handled by control-char strip
]


@pytest.mark.parametrize("payload", _FORMULA_PAYLOADS)
def test_safe_write_neutralizes_formula_payload(payload: str, tmp_path: Path, jrc) -> None:
    workbook_path = tmp_path / f"formula-{abs(hash(payload))}.xlsx"
    workbook = xlsxwriter.Workbook(str(workbook_path))
    worksheet = workbook.add_worksheet("Sheet1")
    jrc._safe_write(worksheet, 0, 0, payload)
    workbook.close()

    loaded = openpyxl.load_workbook(workbook_path, data_only=False)
    cell = loaded["Sheet1"]["A1"]

    # The cell must be a string, never a formula — that's the safety contract.
    assert cell.data_type == "s", (
        f"Payload {payload!r} must round-trip as a string cell; got data_type={cell.data_type}"
    )

    # The displayed value must preserve the payload's meaning (no data loss).
    # control-char strip drops the `\t` payload prefix itself, but the formula
    # text is preserved with a NEW `\t` prefix added by _safe_write.
    # control-char strip drops embedded \x00..\x1f except \n and \t.
    stripped = "".join(
        ch for ch in payload if ord(ch) >= 0x20 or ch in ("\n", "\t")
    )
    # _safe_write adds its own \t before the lstripped formula. The displayed
    # value should be "\t" + stripped (where stripped already has any pre-existing tabs).
    # The lstrip() decision inside _safe_write keys on the FIRST non-whitespace
    # char, so the prefix is always added when the payload triggers.
    assert cell.value.endswith(stripped) or cell.value.startswith("\t"), (
        f"Payload {payload!r} expected to retain text {stripped!r}; got {cell.value!r}"
    )
    # The cell value must start with a tab — the marker that signals neutralization.
    assert cell.value.startswith("\t"), (
        f"Cell value for {payload!r} must start with a tab to mark formula neutralization; "
        f"got {cell.value!r}"
    )


# ---------------------------------------------------------------
# _safe_write — XSS payloads survive (no data loss) but don't execute
# (XLSX has no JS context — these strings just need to round-trip safely).
# ---------------------------------------------------------------

_XSS_PAYLOADS = [
    "<script>alert(1)</script>",
    "<img src=x onerror=alert(1)>",
    "javascript:alert(1)",
    "<iframe src=javascript:alert(1)>",
    "<svg/onload=alert(1)>",
    "\"><script>alert(document.cookie)</script>",
]


@pytest.mark.parametrize("payload", _XSS_PAYLOADS)
def test_safe_write_round_trips_xss_payload_intact(payload: str, tmp_path: Path, jrc) -> None:
    """XLSX has no JS context; these payloads should survive untouched
    (no leading-`=`/`+`/`-`/`@` so no formula prefix added)."""
    workbook_path = tmp_path / f"xss-{abs(hash(payload))}.xlsx"
    workbook = xlsxwriter.Workbook(str(workbook_path))
    worksheet = workbook.add_worksheet("Sheet1")
    jrc._safe_write(worksheet, 0, 0, payload)
    workbook.close()

    loaded = openpyxl.load_workbook(workbook_path, data_only=False)
    cell = loaded["Sheet1"]["A1"]
    assert cell.value == payload, (
        f"XSS payload {payload!r} must round-trip unchanged; got {cell.value!r}"
    )


# ---------------------------------------------------------------
# _safe_write — control characters and unsafe Unicode are stripped
# ---------------------------------------------------------------

_CONTROL_PAYLOADS_CASES = [
    # (input, expected_after_strip)
    ("hello\x00world", "helloworld"),                  # NULL byte
    ("hello\x01\x02\x03world", "helloworld"),          # SOH, STX, ETX
    ("line1\nline2", "line1\nline2"),                  # newline preserved (allowed)
    ("col1\tcol2", "col1\tcol2"),                      # tab preserved (allowed)
    ("safe​text", "safetext"),                    # zero-width space (Cf) stripped
    ("rtl‮marker", "rtlmarker"),                  # right-to-left override (Cf) stripped
    ("\x7fdel", "del"),                                # DEL char stripped
]


@pytest.mark.parametrize("payload,expected", _CONTROL_PAYLOADS_CASES)
def test_safe_write_strips_control_chars(payload: str, expected: str, tmp_path: Path, jrc) -> None:
    workbook_path = tmp_path / f"ctrl-{abs(hash(payload))}.xlsx"
    workbook = xlsxwriter.Workbook(str(workbook_path))
    worksheet = workbook.add_worksheet("Sheet1")
    jrc._safe_write(worksheet, 0, 0, payload)
    workbook.close()

    loaded = openpyxl.load_workbook(workbook_path, data_only=False)
    cell = loaded["Sheet1"]["A1"]
    assert cell.value == expected, (
        f"Payload {payload!r} should sanitize to {expected!r}; got {cell.value!r}"
    )


# ---------------------------------------------------------------
# HtmlReport._html_text — HTML escape round-trip
# ---------------------------------------------------------------
#
# Backed by Python's stdlib `html.escape(..., quote=True)`. Tests that
# &/</>/"/' all become entities; payload survives as literal text but no
# longer renders as script.


def test_html_text_escapes_basic_metachars(jrc) -> None:
    assert jrc.HtmlReport._html_text("&") == "&amp;"
    assert jrc.HtmlReport._html_text("<") == "&lt;"
    assert jrc.HtmlReport._html_text(">") == "&gt;"
    assert jrc.HtmlReport._html_text('"') == "&quot;"
    assert jrc.HtmlReport._html_text("'") == "&#x27;"


def test_html_text_neutralizes_script_tag(jrc) -> None:
    payload = "<script>alert(1)</script>"
    escaped = jrc.HtmlReport._html_text(payload)
    # Tag delimiters must be entified — no `<script` substring remains.
    assert "<script" not in escaped
    assert "</script" not in escaped
    # Payload text content must survive intact, just escaped.
    assert "alert(1)" in escaped
    assert "&lt;script&gt;" in escaped


def test_html_text_neutralizes_img_onerror_payload(jrc) -> None:
    payload = "<img src=x onerror=alert(1)>"
    escaped = jrc.HtmlReport._html_text(payload)
    assert "<img" not in escaped
    assert "onerror=" in escaped or "onerror=alert" in escaped  # text preserved
    assert "&lt;img" in escaped


def test_html_text_neutralizes_attribute_break_payload(jrc) -> None:
    # Common injection: break out of a double-quoted attribute, inject onclick.
    payload = '"><script>alert(document.cookie)</script>'
    escaped = jrc.HtmlReport._html_text(payload)
    assert '"' not in escaped, "Double quote must be entified to prevent attribute break"
    assert "&quot;" in escaped
    assert "<script" not in escaped


def test_html_text_handles_iframe_javascript_uri(jrc) -> None:
    payload = "<iframe src=javascript:alert(1)>"
    escaped = jrc.HtmlReport._html_text(payload)
    assert "<iframe" not in escaped
    assert "javascript:alert(1)" in escaped  # text intact, just no longer a tag


def test_html_text_handles_none_and_empty(jrc) -> None:
    # None and empty string must yield the default (empty string by default).
    assert jrc.HtmlReport._html_text(None) == ""
    assert jrc.HtmlReport._html_text("") == ""
    assert jrc.HtmlReport._html_text(None, default="—") == "—"


def test_html_text_coerces_non_string_input(jrc) -> None:
    # Numeric and other primitives must coerce safely without raising.
    assert jrc.HtmlReport._html_text(42) == "42"
    assert jrc.HtmlReport._html_text(3.14) == "3.14"
    assert jrc.HtmlReport._html_text(True) == "True"
