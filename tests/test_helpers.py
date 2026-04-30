"""Unit tests for helper functions."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import xlsxwriter


# ---------------------------------------------------------------------------
# Helpers for CoreDashboard unit tests
# ---------------------------------------------------------------------------

def _make_core_dashboard(jrc, tmp_path, bridge_data: dict):
    """Build a CoreDashboard backed by a mock bridge returning bridge_data."""
    config = jrc.Config(jrc.Config._WORKSPACE_INIT_DEFAULTS_NAME)

    bridge = MagicMock()
    for attr, value in bridge_data.items():
        getattr(bridge, attr).return_value = value

    wb_path = str(tmp_path / "test.xlsx")
    wb = xlsxwriter.Workbook(wb_path)
    fmts = jrc._build_formats(wb)
    dashboard = jrc.CoreDashboard(config, bridge, wb, fmts)
    return dashboard, wb, wb_path


def _read_sheet_column(wb_path: str, sheet_name: str, col: int) -> list:
    """Return all non-None values in a column (1-indexed) from an xlsx sheet."""
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    ws = wb[sheet_name]
    return [ws.cell(row=r, column=col).value for r in range(1, ws.max_row + 1)
            if ws.cell(row=r, column=col).value is not None]


def _read_summary(wb_path: str, sheet_name: str) -> dict:
    """Return the label→value summary dict for a CoreDashboard sheet.

    Reads col A (label) and col B (value), stopping before any row where col A
    looks like a column header (i.e., where col C is also non-None, indicating a
    multi-column data table has started).
    """
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    ws = wb[sheet_name]
    result: dict = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        value = ws.cell(row=r, column=2).value
        col_c = ws.cell(row=r, column=3).value
        # Once we hit a row where col C is also non-None, we've entered the data table.
        if col_c is not None:
            break
        if label is not None and value is not None:
            result[label] = value
    return result


# ---------------------------------------------------------------------------
# Fail-closed: device compliance — absent stale field
# ---------------------------------------------------------------------------

def test_device_compliance_absent_stale_marked_unknown(jrc, tmp_path) -> None:
    """Devices with no stale field must show 'Unknown', not 'No' (fail-open)."""
    rows = [
        {"name": "Known-Stale", "serial": "AAA", "managed": True,
         "os_version": "15.0", "last_contact": "2020-01-01", "days_since_contact": "1500",
         "stale": True},
        {"name": "Known-Current", "serial": "BBB", "managed": True,
         "os_version": "15.0", "last_contact": "2026-04-01", "days_since_contact": "5",
         "stale": False},
        {"name": "Unknown-Stale", "serial": "CCC", "managed": True,
         "os_version": "15.0", "last_contact": "", "days_since_contact": ""},
        # stale key absent — should NOT be treated as current
    ]
    dashboard, wb, wb_path = _make_core_dashboard(
        jrc, tmp_path, {"device_compliance": rows}
    )
    dashboard._write_device_compliance()
    wb.close()

    stale_col = _read_sheet_column(wb_path, "Device Compliance", col=7)
    assert "Yes" in stale_col, "Confirmed-stale device should show Yes"
    assert "No" in stale_col, "Confirmed-current device should show No"
    assert "Unknown" in stale_col, "Device with absent stale field must show Unknown"
    assert stale_col.count("No") == 1, "Only one device has stale=False"


def test_device_compliance_absent_stale_not_counted_as_stale(jrc, tmp_path) -> None:
    """Device with absent stale field must not inflate confirmed-stale count."""
    rows = [
        {"name": "Current-A", "serial": "A", "managed": True, "stale": False,
         "os_version": "15.0", "last_contact": "2026-04-01", "days_since_contact": "5"},
        # stale key absent — neither confirmed-stale nor confirmed-current
        {"name": "Unknown-B", "serial": "B", "managed": True,
         "os_version": "15.0", "last_contact": "", "days_since_contact": ""},
    ]
    dashboard, wb, wb_path = _make_core_dashboard(
        jrc, tmp_path, {"device_compliance": rows}
    )
    dashboard._write_device_compliance()
    wb.close()

    summary = _read_summary(wb_path, "Device Compliance")

    assert summary.get("Total Devices") == 2
    # stale_count counts only explicit stale=True; absent field must not appear as stale
    assert summary.get("Stale Devices (>30 days)") == 0


# ---------------------------------------------------------------------------
# Fail-closed: platform compliance rules — absent failed field
# ---------------------------------------------------------------------------

def test_platform_compliance_rules_absent_failed_shows_dash(jrc, tmp_path) -> None:
    """Rules with absent 'failed' key must display '—', not '0' (fail-open)."""
    rows = [
        {"rule": "Rule-KnownFail", "passed": 80, "failed": 5,
         "unknown": 0, "devices": 85, "passRate": "94%"},
        {"rule": "Rule-KnownPass", "passed": 90, "failed": 0,
         "unknown": 0, "devices": 90, "passRate": "100%"},
        {"rule": "Rule-MissingData", "passed": 0,
         # "failed" key intentionally absent
         "unknown": 0, "devices": 50, "passRate": ""},
    ]
    dashboard, wb, wb_path = _make_core_dashboard(
        jrc, tmp_path,
        {"compliance_rules": rows},
    )
    dashboard._write_platform_compliance_rules("TBench")
    wb.close()

    # Sheet name: benchmark + "Rules" joined, spaces stripped by _excel_sheet_name
    wb_loaded = openpyxl.load_workbook(wb_path, data_only=False)
    sheet_name = wb_loaded.sheetnames[0]
    ws = wb_loaded[sheet_name]
    # Column 3 (C) is "Failed" in the detail table
    failed_values = [ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1)]
    assert "—" in failed_values, "Rule with absent failed field must display '—'"
    # Explicit 0 is valid; absent field must NOT produce a 0
    explicit_zeros = sum(1 for v in failed_values if v == 0)
    assert explicit_zeros <= 1, "Only the known-0 rule should produce a 0 in the Failed column"


def test_platform_compliance_rules_missing_data_count_in_summary(jrc, tmp_path) -> None:
    """'Rules with Missing Data' summary count must be non-zero when failed key is absent."""
    rows = [
        {"rule": "RuleA", "passed": 10, "failed": 0, "unknown": 0, "devices": 10,
         "passRate": "100%"},
        {"rule": "RuleB", "passed": 5,
         # failed key absent
         "unknown": 0, "devices": 5, "passRate": ""},
    ]
    dashboard, wb, wb_path = _make_core_dashboard(
        jrc, tmp_path, {"compliance_rules": rows}
    )
    dashboard._write_platform_compliance_rules("MBench")
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path, data_only=False)
    sheet_name = wb_loaded.sheetnames[0]
    summary = _read_summary(wb_path, sheet_name)
    assert summary.get("Rules with Missing Data") == 1


# ---------------------------------------------------------------------------
# Fail-closed: platform compliance devices — absent rulesFailed field
# ---------------------------------------------------------------------------

def test_platform_compliance_devices_absent_rules_failed_shows_dash(jrc, tmp_path) -> None:
    """Devices with absent 'rulesFailed' key must display '—', not '0' (fail-open)."""
    rows = [
        {"device": "MacA", "deviceId": "1", "rulesFailed": 3,
         "rulesPassed": 7, "compliance": "70%"},
        {"device": "MacB", "deviceId": "2", "rulesFailed": 0,
         "rulesPassed": 10, "compliance": "100%"},
        {"device": "MacC", "deviceId": "3",
         # rulesFailed key intentionally absent
         "rulesPassed": 0, "compliance": ""},
    ]
    dashboard, wb, wb_path = _make_core_dashboard(
        jrc, tmp_path, {"compliance_devices": rows}
    )
    dashboard._write_platform_compliance_devices("TBench2")
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path, data_only=False)
    sheet_name = wb_loaded.sheetnames[0]
    ws = wb_loaded[sheet_name]
    # Column 3 is "Rules Failed" in the detail table
    rules_failed_values = [ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1)]
    assert "—" in rules_failed_values, "Device with absent rulesFailed must display '—'"


def test_platform_compliance_devices_unknown_status_count_in_summary(jrc, tmp_path) -> None:
    """'Devices with Unknown Status' must reflect devices missing rulesFailed."""
    rows = [
        {"device": "MacA", "deviceId": "1", "rulesFailed": 2, "rulesPassed": 8,
         "compliance": "80%"},
        {"device": "MacB", "deviceId": "2",
         # rulesFailed absent
         "rulesPassed": 0, "compliance": ""},
        {"device": "MacC", "deviceId": "3",
         # rulesFailed absent
         "rulesPassed": 0, "compliance": ""},
    ]
    dashboard, wb, wb_path = _make_core_dashboard(
        jrc, tmp_path, {"compliance_devices": rows}
    )
    dashboard._write_platform_compliance_devices("MBench2")
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path, data_only=False)
    sheet_name = wb_loaded.sheetnames[0]
    summary = _read_summary(wb_path, sheet_name)
    assert summary.get("Devices with Unknown Status") == 2


# ---------------------------------------------------------------------------
# Original helper tests (unchanged)
# ---------------------------------------------------------------------------

def test_parse_manager_extracts_cn_from_dn(jrc) -> None:
    raw = r"CN=SMITH\, JOHN,OU=People,DC=example,DC=com"
    assert jrc._parse_manager(raw) == "Smith, John"


def test_to_int_and_to_bool_handle_common_edge_cases(jrc) -> None:
    assert jrc._to_int("42.9") == 42
    assert jrc._to_int("abc", default=7) == 7
    assert jrc._to_bool("YES") is True
    assert jrc._to_bool("0") is False


def test_safe_write_sanitizes_formula_control_chars_and_inf(tmp_path: Path, jrc) -> None:
    workbook_path = tmp_path / "safe-write.xlsx"
    workbook = xlsxwriter.Workbook(str(workbook_path))
    worksheet = workbook.add_worksheet("Sheet1")

    jrc._safe_write(worksheet, 0, 0, "=SUM(1,2)")
    jrc._safe_write(worksheet, 1, 0, "hello\x00world")
    jrc._safe_write(worksheet, 2, 0, float("inf"))
    jrc._safe_write(worksheet, 3, 0, None)
    workbook.close()

    loaded = openpyxl.load_workbook(workbook_path, data_only=False)
    sheet = loaded["Sheet1"]
    assert sheet["A1"].value == "\t=SUM(1,2)"
    assert sheet["A1"].data_type == "s"
    assert sheet["A2"].value == "helloworld"
    assert sheet["A3"].value == 0
    assert sheet["A4"].value is None


def test_semantic_warnings_flags_managed_as_manager(jrc) -> None:
    config = jrc.Config(jrc.Config._WORKSPACE_INIT_DEFAULTS_NAME)
    config._data["columns"]["manager"] = "Managed"
    df = jrc.pd.DataFrame({"Managed": ["Managed", "Unmanaged"]})
    warnings = jrc._semantic_warnings(config, df)
    assert any("management-state column" in warning for warning in warnings)


def test_inventory_ea_results_are_optional(jrc) -> None:
    class FailingBridge:
        def ea_results_report(self, include_all: bool = True):
            raise RuntimeError("jamf-cli failed")

    ea_columns, unmatched, error = jrc._apply_inventory_ea_results(FailingBridge(), {})

    assert ea_columns == set()
    assert unmatched == 0
    assert "jamf-cli failed" in error


def test_collect_command_plan_includes_cli_only_surfaces(jrc) -> None:
    class FakeBridge:
        def __getattr__(self, _name: str):
            return lambda *args, **kwargs: None

    config = jrc.Config(jrc.Config._WORKSPACE_INIT_DEFAULTS_NAME)
    labels = [
        label
        for label, _command in jrc._collect_jamf_cli_commands(config, FakeBridge(), True)
    ]

    assert "Computer Inventory" in labels
    assert "App Status" in labels
    assert "Update Failures" in labels
    assert "Package Lifecycle" in labels
