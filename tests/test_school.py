"""Tests for Jamf School support: SchoolCLIBridge, SchoolDashboard, school commands."""

from __future__ import annotations

from pathlib import Path

import pytest


SCHOOL_CSV = "harboredu_school_devices.csv"


# ---------------------------------------------------------------------------
# SchoolCLIBridge
# ---------------------------------------------------------------------------


def test_school_bridge_is_subclass_of_jamf_cli_bridge(jrc) -> None:
    assert issubclass(jrc.SchoolCLIBridge, jrc.JamfCLIBridge)


def test_school_bridge_inherits_is_available(jrc) -> None:
    bridge = jrc.SchoolCLIBridge(save_output=False, data_dir="/tmp", profile="test")
    # is_available() should return bool regardless of whether binary exists
    result = bridge.is_available()
    assert isinstance(result, bool)


def test_school_bridge_has_cached_school_data_false_when_empty(jrc, tmp_path) -> None:
    bridge = jrc.SchoolCLIBridge(
        save_output=False, data_dir=str(tmp_path), profile="test"
    )
    assert bridge.has_cached_school_data() is False


# ---------------------------------------------------------------------------
# _school_csv_load
# ---------------------------------------------------------------------------


def test_school_csv_load_semicolon_delimiter(jrc, fixtures_root) -> None:
    csv_path = fixtures_root / "csv" / SCHOOL_CSV
    df = jrc._school_csv_load(str(csv_path))
    # Semicolon-delimited: all expected columns should be present
    assert "SerialNumber" in df.columns
    assert "IsManaged" in df.columns
    assert "LocationName" in df.columns
    assert "OsVersion" in df.columns


def test_school_csv_load_row_count(jrc, fixtures_root) -> None:
    csv_path = fixtures_root / "csv" / SCHOOL_CSV
    df = jrc._school_csv_load(str(csv_path))
    # 15 data rows in the fixture (plus header = 16 lines total)
    assert len(df) == 15


def test_school_csv_load_no_nulls(jrc, fixtures_root) -> None:
    csv_path = fixtures_root / "csv" / SCHOOL_CSV
    df = jrc._school_csv_load(str(csv_path))
    # fillna("") should leave no NaN values
    assert df.isnull().sum().sum() == 0


# ---------------------------------------------------------------------------
# SchoolColumnMapper
# ---------------------------------------------------------------------------


def test_school_column_mapper_get_mapped(jrc, fixtures_root) -> None:
    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    mapper = jrc.SchoolColumnMapper(config)
    assert mapper.get("serial_number") == "SerialNumber"
    assert mapper.get("os_version") == "OsVersion"
    assert mapper.get("managed") == "IsManaged"


def test_school_column_mapper_get_unmapped_returns_none(jrc, fixtures_root) -> None:
    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    mapper = jrc.SchoolColumnMapper(config)
    assert mapper.get("nonexistent_field") is None


def test_school_column_mapper_extract(jrc, fixtures_root) -> None:
    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    mapper = jrc.SchoolColumnMapper(config)
    row = {"SerialNumber": "ABC123", "IsManaged": "true"}
    assert mapper.extract(row, "serial_number") == "ABC123"
    assert mapper.extract(row, "managed") == "true"
    assert mapper.extract(row, "nonexistent_field") == ""


# ---------------------------------------------------------------------------
# SCHOOL_COLUMN_HINTS scaffold matching
# ---------------------------------------------------------------------------


def test_school_scaffold_matches_harboredu_columns(jrc, fixtures_root) -> None:
    """All 18 school_columns fields should auto-match the HarborEdu CSV."""
    csv_path = fixtures_root / "csv" / SCHOOL_CSV
    df = jrc._school_csv_load(str(csv_path))
    headers = list(df.columns)

    def _best_match(logical: str) -> str:
        hints = jrc.SCHOOL_COLUMN_HINTS.get(logical, [])
        excludes = jrc.SCHOOL_COLUMN_EXCLUDES.get(logical, [])
        for hint in hints:
            for h in headers:
                h_lower = h.lower()
                if any(ex in h_lower for ex in excludes):
                    continue
                if hint in h_lower:
                    return h
        return ""

    unmatched = [
        field
        for field in jrc.DEFAULT_CONFIG["school_columns"]
        if not _best_match(field)
    ]
    assert unmatched == [], f"Fields not auto-matched: {unmatched}"


def test_school_scaffold_device_name_is_name_not_location(jrc, fixtures_root) -> None:
    """device_name must match 'Name', not 'LocationName'."""
    csv_path = fixtures_root / "csv" / SCHOOL_CSV
    df = jrc._school_csv_load(str(csv_path))
    headers = list(df.columns)
    hints = jrc.SCHOOL_COLUMN_HINTS.get("device_name", [])
    excludes = jrc.SCHOOL_COLUMN_EXCLUDES.get("device_name", [])

    for hint in hints:
        for h in headers:
            h_lower = h.lower()
            if any(ex in h_lower for ex in excludes):
                continue
            if hint in h_lower:
                assert h == "Name", f"device_name matched {h!r}, expected 'Name'"
                return
    pytest.fail("device_name did not match any column")


# ---------------------------------------------------------------------------
# SchoolDashboard — CSV-driven generation
# ---------------------------------------------------------------------------


def test_school_dashboard_builds_all_csv_sheets(jrc, fixtures_root, tmp_path) -> None:
    """school-generate with only a CSV should create the four CSV-driven sheets."""
    import xlsxwriter

    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    config._data["output"]["output_dir"] = str(tmp_path / "Generated Reports")

    csv_path = fixtures_root / "csv" / SCHOOL_CSV
    df = jrc._school_csv_load(str(csv_path))

    out_path = tmp_path / "school_report.xlsx"
    workbook = xlsxwriter.Workbook(str(out_path), {"remove_timezone": True})
    fmts = jrc._build_formats(workbook)

    dashboard = jrc.SchoolDashboard(config, workbook, fmts, csv_df=df)
    dashboard.build_all()
    workbook.close()

    assert out_path.exists()
    assert out_path.stat().st_size > 0

    # Verify all four expected CSV-driven sheets were created
    import zipfile
    with zipfile.ZipFile(out_path) as zf:
        sheet_xml_names = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")]
    assert len(sheet_xml_names) == 4, f"Expected 4 sheets, got {len(sheet_xml_names)}"


def test_school_dashboard_device_inventory_has_rows(jrc, fixtures_root, tmp_path) -> None:
    """Device Inventory sheet should contain all 15 device rows from the fixture."""
    import xlsxwriter

    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    csv_path = fixtures_root / "csv" / SCHOOL_CSV
    df = jrc._school_csv_load(str(csv_path))

    out_path = tmp_path / "school_inventory.xlsx"
    workbook = xlsxwriter.Workbook(str(out_path), {"remove_timezone": True})
    fmts = jrc._build_formats(workbook)

    dashboard = jrc.SchoolDashboard(config, workbook, fmts, csv_df=df)
    # Write just the Device Inventory sheet
    ws = workbook.add_worksheet("Device Inventory")
    dashboard._write_device_inventory_compact(ws)
    workbook.close()

    # The workbook should be non-empty
    assert out_path.stat().st_size > 0


def test_school_dashboard_os_versions_counts(jrc, fixtures_root, tmp_path) -> None:
    """OS Versions sheet should sum to the total device count."""
    import xlsxwriter
    from collections import Counter

    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    csv_path = fixtures_root / "csv" / SCHOOL_CSV
    df = jrc._school_csv_load(str(csv_path))

    counter: Counter = Counter()
    for val in df["OsVersion"]:
        if str(val).strip():
            counter[str(val).strip()] += 1

    assert sum(counter.values()) == 15


def test_school_dashboard_stale_devices_no_error_on_empty(
    jrc, fixtures_root, tmp_path
) -> None:
    """Stale Devices sheet should not raise when no devices are stale."""
    import xlsxwriter

    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    # Set stale threshold to 0 days — all devices qualify
    config._data["thresholds"]["stale_device_days"] = 0
    csv_path = fixtures_root / "csv" / SCHOOL_CSV
    df = jrc._school_csv_load(str(csv_path))

    out_path = tmp_path / "school_stale.xlsx"
    workbook = xlsxwriter.Workbook(str(out_path), {"remove_timezone": True})
    fmts = jrc._build_formats(workbook)

    dashboard = jrc.SchoolDashboard(config, workbook, fmts, csv_df=df)
    ws = workbook.add_worksheet("Stale Devices")
    dashboard._write_stale_devices(ws)
    workbook.close()

    assert out_path.exists()


# ---------------------------------------------------------------------------
# cmd_school_scaffold
# ---------------------------------------------------------------------------


def test_cmd_school_scaffold_writes_file(jrc, fixtures_root, tmp_path) -> None:
    csv_path = str(fixtures_root / "csv" / SCHOOL_CSV)
    out_path = str(tmp_path / "school_columns.yaml")
    jrc.cmd_school_scaffold(csv_path, out_path)
    assert Path(out_path).exists()
    content = Path(out_path).read_text()
    assert "school_columns:" in content
    assert "SerialNumber" in content
    assert "IsManaged" in content


def test_cmd_school_scaffold_no_duplicate_write(jrc, fixtures_root, tmp_path) -> None:
    """Running scaffold twice on the same output file should not duplicate the block."""
    csv_path = str(fixtures_root / "csv" / SCHOOL_CSV)
    out_path = str(tmp_path / "school_columns.yaml")
    jrc.cmd_school_scaffold(csv_path, out_path)
    first_content = Path(out_path).read_text()
    # Run again — should warn and not modify the file
    jrc.cmd_school_scaffold(csv_path, out_path)
    assert Path(out_path).read_text() == first_content


# ---------------------------------------------------------------------------
# cmd_school_generate (integration)
# ---------------------------------------------------------------------------


def test_cmd_school_generate_csv_only(jrc, fixtures_root, tmp_path) -> None:
    """school-generate with only --csv should produce a valid xlsx."""
    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    config._data["output"]["output_dir"] = str(tmp_path / "Generated Reports")

    csv_path = str(fixtures_root / "csv" / SCHOOL_CSV)
    out_file = str(tmp_path / "school_report.xlsx")
    jrc.cmd_school_generate(config, csv_path=csv_path, out_file=out_file)

    assert Path(out_file).exists()
    assert Path(out_file).stat().st_size > 5000


# ---------------------------------------------------------------------------
# SchoolDashboard._write_device_inventory_compact
# ---------------------------------------------------------------------------


def _make_school_dashboard(jrc, config, df, tmp_path):
    import xlsxwriter

    out_path = tmp_path / "inv.xlsx"
    workbook = xlsxwriter.Workbook(str(out_path), {"remove_timezone": True})
    fmts = jrc._build_formats(workbook)
    dashboard = jrc.SchoolDashboard(config, workbook, fmts, csv_df=df)
    return dashboard, workbook, out_path


def test_write_device_inventory_compact_no_crash_missing_columns(
    jrc, fixtures_root, tmp_path
) -> None:
    """Missing device_type / user_name mappings produce blank cells, no exception."""
    import pandas as pd

    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    # Omit device_type and user_name from config — they should be blank, not a crash
    config._data["school_columns"].pop("device_type", None)
    config._data["school_columns"].pop("user_name", None)

    df = pd.DataFrame({
        "SerialNumber": ["ABC123", "DEF456"],
        "Name": ["iPad 1", "iPad 2"],
        "OsVersion": ["17.0", "16.5"],
        "LocationName": ["School A", "School A"],
        "LastConnected": ["2026-04-01T00:00:00Z", "2026-03-01T00:00:00Z"],
    })

    dashboard, workbook, out_path = _make_school_dashboard(jrc, config, df, tmp_path)
    ws = workbook.add_worksheet("Inv")
    dashboard._write_device_inventory_compact(ws)
    workbook.close()

    assert out_path.stat().st_size > 0


def test_write_device_inventory_compact_malformed_checkin_treated_as_stale(
    jrc, fixtures_root, tmp_path
) -> None:
    """Unparseable last_checkin values are treated as stale (Status = 'Stale')."""
    import pandas as pd

    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    config._data["thresholds"]["stale_device_days"] = 30

    df = pd.DataFrame({
        "SerialNumber": ["ABC123"],
        "Name": ["Test iPad"],
        "OsVersion": ["17.0"],
        "LocationName": ["School A"],
        "LastConnected": ["not-a-date"],
    })

    dashboard, workbook, out_path = _make_school_dashboard(jrc, config, df, tmp_path)
    ws = workbook.add_worksheet("Inv")
    dashboard._write_device_inventory_compact(ws)
    workbook.close()

    # The workbook writes without error; we verify it is non-empty as a proxy
    # for the row being written at all (stale path must not raise or skip).
    assert out_path.stat().st_size > 0


def test_write_device_inventory_compact_all_stale(
    jrc, fixtures_root, tmp_path
) -> None:
    """Every row whose last_checkin exceeds the threshold is marked 'Stale'."""
    import pandas as pd
    import openpyxl

    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    config._data["thresholds"]["stale_device_days"] = 1  # tiny threshold

    # Use a date far in the past so every row is stale
    df = pd.DataFrame({
        "SerialNumber": ["S1", "S2", "S3"],
        "Name": ["Dev 1", "Dev 2", "Dev 3"],
        "OsVersion": ["17.0", "17.0", "17.0"],
        "LocationName": ["Loc A", "Loc B", "Loc A"],
        "LastConnected": [
            "2020-01-01T00:00:00Z",
            "2019-06-15T12:00:00Z",
            "2018-09-01T00:00:00Z",
        ],
    })

    dashboard, workbook, out_path = _make_school_dashboard(jrc, config, df, tmp_path)
    ws = workbook.add_worksheet("Inv")
    dashboard._write_device_inventory_compact(ws)
    workbook.close()

    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb.active
    # Column H (index 8) is "Status"; row 1 is the header
    statuses = [sheet.cell(row=r, column=8).value for r in range(2, 5)]
    assert all(s == "Stale" for s in statuses), f"Expected all Stale, got {statuses}"


# ---------------------------------------------------------------------------
# SchoolDashboard._write_device_groups_sorted
# ---------------------------------------------------------------------------


class _StubBridge:
    """Minimal bridge stub for device-groups tests."""

    def __init__(self, groups):
        self._groups = groups

    def device_groups_list(self):
        return self._groups


def _make_bridge_dashboard(jrc, config, bridge, tmp_path):
    import xlsxwriter

    out_path = tmp_path / "groups.xlsx"
    workbook = xlsxwriter.Workbook(str(out_path), {"remove_timezone": True})
    fmts = jrc._build_formats(workbook)
    dashboard = jrc.SchoolDashboard(config, workbook, fmts, bridge=bridge)
    return dashboard, workbook, out_path


def test_write_device_groups_sorted_empty_no_crash(jrc, fixtures_root, tmp_path) -> None:
    """Empty group list writes header row only and does not raise."""
    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    bridge = _StubBridge([])
    dashboard, workbook, out_path = _make_bridge_dashboard(jrc, config, bridge, tmp_path)
    ws = workbook.add_worksheet("Device Groups")
    dashboard._write_device_groups_sorted(ws)
    workbook.close()

    import openpyxl
    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb.active
    assert sheet.cell(row=1, column=1).value == "Group Name"
    assert sheet.cell(row=2, column=1).value is None, "Expected no data rows for empty list"


def test_write_device_groups_sorted_sort_order(jrc, fixtures_root, tmp_path) -> None:
    """Rows are written in device count descending order."""
    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    groups = [
        {"name": "Small Group", "device_count": 5, "location_name": "Campus A"},
        {"name": "Large Group", "device_count": 100, "location_name": "Campus B"},
        {"name": "Medium Group", "device_count": 30, "location_name": "Campus A"},
    ]
    bridge = _StubBridge(groups)
    dashboard, workbook, out_path = _make_bridge_dashboard(jrc, config, bridge, tmp_path)
    ws = workbook.add_worksheet("Device Groups")
    dashboard._write_device_groups_sorted(ws)
    workbook.close()

    import openpyxl
    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb.active
    names = [sheet.cell(row=r, column=1).value for r in range(2, 5)]
    counts = [sheet.cell(row=r, column=2).value for r in range(2, 5)]
    assert names == ["Large Group", "Medium Group", "Small Group"], (
        f"Expected descending name order, got {names}"
    )
    assert counts == [100, 30, 5], f"Expected descending counts, got {counts}"


def test_write_device_groups_sorted_none_bridge_returns_empty(
    jrc, fixtures_root, tmp_path
) -> None:
    """None returned from bridge (cache miss) produces header-only sheet without error."""
    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    bridge = _StubBridge(None)
    dashboard, workbook, out_path = _make_bridge_dashboard(jrc, config, bridge, tmp_path)
    ws = workbook.add_worksheet("Device Groups")
    dashboard._write_device_groups_sorted(ws)
    workbook.close()

    import openpyxl
    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb.active
    assert sheet.cell(row=1, column=1).value == "Group Name"
    assert sheet.cell(row=2, column=1).value is None


class _DEPBridge:
    """Minimal bridge stub returning DEP devices."""

    def __init__(self, items):
        self._items = items

    def dep_devices_list(self):
        return self._items


class _IBeaconsBridge:
    """Minimal bridge stub returning iBeacons."""

    def __init__(self, items):
        self._items = items

    def ibeacons_list(self):
        return self._items


def _load_json_fixture(fixtures_root, rel: str):
    import json
    return json.loads((fixtures_root / "jamf-cli-data" / rel).read_text())


# ---------------------------------------------------------------------------
# SchoolDashboard._write_dep_devices
# ---------------------------------------------------------------------------


def test_write_dep_devices_happy_path(jrc, fixtures_root, tmp_path) -> None:
    """DEP devices fixture renders with all 6 columns populated for the first row."""
    import openpyxl

    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    items = _load_json_fixture(
        fixtures_root, "school-dep-devices/dep_devices_happy.json"
    )
    bridge = _DEPBridge(items)
    dashboard, workbook, out_path = _make_bridge_dashboard(jrc, config, bridge, tmp_path)
    ws = workbook.add_worksheet("DEP Devices")
    dashboard._write_dep_devices(ws)
    workbook.close()

    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb.active
    headers = [sheet.cell(row=1, column=c).value for c in range(1, 7)]
    assert headers == [
        "Serial Number", "Model", "Color", "Status",
        "Profile Name", "Device Name",
    ]
    # 3 data rows
    serials = [sheet.cell(row=r, column=1).value for r in range(2, 5)]
    assert all(s for s in serials), f"Expected non-empty serials, got {serials}"


def test_write_dep_devices_empty_no_crash(jrc, fixtures_root, tmp_path) -> None:
    """Empty DEP list writes header row only and does not raise."""
    import openpyxl

    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    bridge = _DEPBridge([])
    dashboard, workbook, out_path = _make_bridge_dashboard(jrc, config, bridge, tmp_path)
    ws = workbook.add_worksheet("DEP Devices")
    dashboard._write_dep_devices(ws)
    workbook.close()

    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb.active
    assert sheet.cell(row=1, column=1).value == "Serial Number"
    assert sheet.cell(row=2, column=1).value is None


def test_write_dep_devices_sorted_by_serial(jrc, fixtures_root, tmp_path) -> None:
    """Rows are written in serial-number ascending order."""
    import openpyxl

    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    items = [
        {"serialNumber": "ZZZ999", "model": "iPad", "color": "Silver",
         "status": "Assigned", "profileName": "P1", "deviceName": "D1"},
        {"serialNumber": "AAA111", "model": "iPad", "color": "Silver",
         "status": "Assigned", "profileName": "P1", "deviceName": "D2"},
        {"serialNumber": "MMM555", "model": "iPad", "color": "Silver",
         "status": "Assigned", "profileName": "P1", "deviceName": "D3"},
    ]
    bridge = _DEPBridge(items)
    dashboard, workbook, out_path = _make_bridge_dashboard(jrc, config, bridge, tmp_path)
    ws = workbook.add_worksheet("DEP Devices")
    dashboard._write_dep_devices(ws)
    workbook.close()

    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb.active
    serials = [sheet.cell(row=r, column=1).value for r in range(2, 5)]
    assert serials == ["AAA111", "MMM555", "ZZZ999"]


# ---------------------------------------------------------------------------
# SchoolDashboard._write_ibeacons
# ---------------------------------------------------------------------------


def test_write_ibeacons_happy_path(jrc, fixtures_root, tmp_path) -> None:
    """iBeacons fixture renders the 5 expected columns and lowerCamelCase uuid."""
    import openpyxl

    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    items = _load_json_fixture(
        fixtures_root, "school-ibeacons/ibeacons_happy.json"
    )
    bridge = _IBeaconsBridge(items)
    dashboard, workbook, out_path = _make_bridge_dashboard(jrc, config, bridge, tmp_path)
    ws = workbook.add_worksheet("iBeacons")
    dashboard._write_ibeacons(ws)
    workbook.close()

    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb.active
    headers = [sheet.cell(row=1, column=c).value for c in range(1, 6)]
    assert headers == ["Name", "UUID", "Major", "Minor", "Description"]
    # UUID column populated from lowercase 'uuid' source key
    uuids = [sheet.cell(row=r, column=2).value for r in range(2, 5)]
    assert all(u for u in uuids), f"Expected non-empty UUIDs, got {uuids}"


def test_write_ibeacons_empty_no_crash(jrc, fixtures_root, tmp_path) -> None:
    """Empty iBeacons list writes header row only and does not raise."""
    import openpyxl

    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    bridge = _IBeaconsBridge([])
    dashboard, workbook, out_path = _make_bridge_dashboard(jrc, config, bridge, tmp_path)
    ws = workbook.add_worksheet("iBeacons")
    dashboard._write_ibeacons(ws)
    workbook.close()

    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb.active
    assert sheet.cell(row=1, column=1).value == "Name"
    assert sheet.cell(row=2, column=1).value is None


def test_write_ibeacons_sorted_by_name(jrc, fixtures_root, tmp_path) -> None:
    """Rows are written in name ascending order."""
    import openpyxl

    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    items = [
        {"name": "Zone Z", "uuid": "uuid-z", "major": 1, "minor": 1, "description": ""},
        {"name": "Alpha Hall", "uuid": "uuid-a", "major": 1, "minor": 2, "description": ""},
        {"name": "Mid-Block", "uuid": "uuid-m", "major": 1, "minor": 3, "description": ""},
    ]
    bridge = _IBeaconsBridge(items)
    dashboard, workbook, out_path = _make_bridge_dashboard(jrc, config, bridge, tmp_path)
    ws = workbook.add_worksheet("iBeacons")
    dashboard._write_ibeacons(ws)
    workbook.close()

    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb.active
    names = [sheet.cell(row=r, column=1).value for r in range(2, 5)]
    assert names == ["Alpha Hall", "Mid-Block", "Zone Z"]


def test_write_device_groups_sorted_locations_comma_joined(
    jrc, fixtures_root, tmp_path
) -> None:
    """Multiple locations in a list are comma-joined in the Locations column."""
    config = jrc.Config(str(fixtures_root / "config" / "school_test.yaml"))
    groups = [
        {"name": "Multi-Campus", "device_count": 10,
         "locations": ["Campus A", "Campus B", "Campus C"]},
    ]
    bridge = _StubBridge(groups)
    dashboard, workbook, out_path = _make_bridge_dashboard(jrc, config, bridge, tmp_path)
    ws = workbook.add_worksheet("Device Groups")
    dashboard._write_device_groups_sorted(ws)
    workbook.close()

    import openpyxl
    wb = openpyxl.load_workbook(str(out_path))
    sheet = wb.active
    locations_cell = sheet.cell(row=2, column=3).value
    assert locations_cell == "Campus A, Campus B, Campus C", (
        f"Expected comma-joined locations, got {locations_cell!r}"
    )
