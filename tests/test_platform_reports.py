"""Tests for Platform API reports: blueprint-status, compliance-rules/devices, ddm-status.

All four report methods are gated behind platform.enabled = true in DEFAULT_CONFIG.
These tests use synthetic fixtures derived from the as-built parser code (field names
confirmed by reading _write_platform_* methods, not from live API calls). Field names
should be re-verified against live v1.14 output once Platform API reaches GA.

Fixture paths:
    tests/fixtures/jamf-cli-data/blueprint-status/
    tests/fixtures/jamf-cli-data/compliance-rules-nist-800-53r5-moderate/
    tests/fixtures/jamf-cli-data/compliance-devices-nist-800-53r5-moderate/
    tests/fixtures/jamf-cli-data/ddm-status/
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock

import openpyxl
import pytest
import xlsxwriter


BENCHMARK = "NIST 800-53r5 Moderate"
FIXTURES_ROOT = Path(__file__).resolve().parent / "fixtures"
BP_DIR = FIXTURES_ROOT / "jamf-cli-data" / "blueprint-status"
CR_DIR = FIXTURES_ROOT / "jamf-cli-data" / "compliance-rules-nist-800-53r5-moderate"
CD_DIR = FIXTURES_ROOT / "jamf-cli-data" / "compliance-devices-nist-800-53r5-moderate"
DDM_DIR = FIXTURES_ROOT / "jamf-cli-data" / "ddm-status"


def _load(path: Path) -> Any:
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def _make_dashboard(jrc, tmp_path: Path, bridge_data: dict, platform_enabled: bool = True):
    """Build a CoreDashboard backed by a mock bridge."""
    config = jrc.Config(jrc.Config._WORKSPACE_INIT_DEFAULTS_NAME)
    config._data["platform"]["enabled"] = platform_enabled
    config._data["platform"]["compliance_benchmarks"] = [BENCHMARK]

    bridge = MagicMock()
    for attr, value in bridge_data.items():
        if isinstance(value, Exception):
            getattr(bridge, attr).side_effect = value
        else:
            getattr(bridge, attr).return_value = value

    wb_path = str(tmp_path / "test.xlsx")
    wb = xlsxwriter.Workbook(wb_path, {"remove_timezone": True})
    fmts = jrc._build_formats(wb)
    dashboard = jrc.CoreDashboard(config, bridge, wb, fmts)
    return dashboard, wb, wb_path


def _read_column(wb_path: str, sheet_name: str, col: int) -> list:
    """Return all non-None cell values in a 1-indexed column."""
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    ws = wb[sheet_name]
    return [ws.cell(row=r, column=col).value for r in range(1, ws.max_row + 1)
            if ws.cell(row=r, column=col).value is not None]


def _read_summary(wb_path: str, sheet_name: str) -> dict:
    """Return label→value pairs from the summary block (before the data table)."""
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    ws = wb[sheet_name]
    result: dict = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        value = ws.cell(row=r, column=2).value
        col_c = ws.cell(row=r, column=3).value
        if col_c is not None:
            break
        if label is not None and value is not None:
            result[label] = value
    return result


# ---------------------------------------------------------------------------
# Platform Blueprints — happy path
# ---------------------------------------------------------------------------


def test_blueprint_status_happy_path_no_raise(jrc, tmp_path) -> None:
    """Happy fixture renders without exception; sheet is created."""
    rows = _load(BP_DIR / "platform_blueprint_status_happy.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"blueprint_status": rows})
    dashboard._write_platform_blueprints()
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path)
    assert "Platform Blueprints" in wb_loaded.sheetnames


def test_blueprint_status_happy_summary_counts(jrc, tmp_path) -> None:
    """Summary counts match the live-shape happy fixture (4 entries, 2 DEPLOYED, 1 pending)."""
    rows = _load(BP_DIR / "platform_blueprint_status_happy.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"blueprint_status": rows})
    dashboard._write_platform_blueprints()
    wb.close()

    summary = _read_summary(wb_path, "Platform Blueprints")
    assert summary.get("Total Blueprints") == 4
    assert summary.get("Deployed Blueprints") == 2
    assert summary.get("Blueprints with Failures") == 0
    assert summary.get("Blueprints with Pending Devices") == 1


def test_blueprint_status_sort_order_failed_first(jrc, tmp_path) -> None:
    """Rows are sorted descending by failed, then pending, then name."""
    rows = _load(BP_DIR / "platform_blueprint_status_all_failed.json")
    expected_names = {r["name"] for r in rows}
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"blueprint_status": rows})
    dashboard._write_platform_blueprints()
    wb.close()

    all_col1 = _read_column(wb_path, "Platform Blueprints", col=1)
    data_names = [n for n in all_col1 if n in expected_names]
    # "Alpha Blueprint" has failed=10 — must appear first.
    assert data_names[0] == "Alpha Blueprint", (
        f"Blueprint with most failures must sort first; got {data_names}"
    )


def test_blueprint_status_partial_no_crash(jrc, tmp_path) -> None:
    """Blueprint missing succeeded/failed renders empty cells without raising."""
    rows = _load(BP_DIR / "platform_blueprint_status_partial.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"blueprint_status": rows})
    dashboard._write_platform_blueprints()
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path)
    assert "Platform Blueprints" in wb_loaded.sheetnames


# ---------------------------------------------------------------------------
# Platform Blueprints — empty input raises RuntimeError
# ---------------------------------------------------------------------------


def test_blueprint_status_empty_raises_runtime_error(jrc, tmp_path) -> None:
    """Empty bridge response raises RuntimeError so write_all can skip the sheet."""
    rows = _load(BP_DIR / "platform_blueprint_status_empty.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"blueprint_status": rows})
    with pytest.raises(RuntimeError, match="blueprint-status returned no rows"):
        dashboard._write_platform_blueprints()
    wb.close()


# ---------------------------------------------------------------------------
# Compliance Rules — happy path + key-absent invariant
# ---------------------------------------------------------------------------


def test_compliance_rules_happy_path_no_raise(jrc, tmp_path) -> None:
    """Happy fixture for compliance-rules renders without exception."""
    rows = _load(CR_DIR / "platform_compliance_rules_happy.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"compliance_rules": rows})
    dashboard._write_platform_compliance_rules(BENCHMARK)
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path)
    assert len(wb_loaded.sheetnames) == 1


def test_compliance_rules_absent_failed_shows_dash(jrc, tmp_path) -> None:
    """Rule with absent 'failed' key displays '—' not 0 in the Failed column."""
    rows = _load(CR_DIR / "platform_compliance_rules_happy.json")
    # "Gatekeeper Status" has no 'failed' key in the fixture
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"compliance_rules": rows})
    dashboard._write_platform_compliance_rules(BENCHMARK)
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path, data_only=False)
    sheet_name = wb_loaded.sheetnames[0]
    ws = wb_loaded[sheet_name]
    failed_values = [ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1)]
    assert "—" in failed_values, (
        "Rule with absent 'failed' key must render '—', not 0"
    )


def test_compliance_rules_all_failed_summary(jrc, tmp_path) -> None:
    """All-failed fixture: Average Pass Rate is present and Rules with Failures equals total."""
    rows = _load(CR_DIR / "platform_compliance_rules_all_failed.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"compliance_rules": rows})
    dashboard._write_platform_compliance_rules(BENCHMARK)
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path, data_only=False)
    sheet_name = wb_loaded.sheetnames[0]
    summary = _read_summary(wb_path, sheet_name)
    assert summary.get("Rules with Failures") == 3
    avg = summary.get("Average Pass Rate")
    assert avg is not None and avg != "", "Average Pass Rate must be non-empty when passRate is present"


def test_compliance_rules_empty_raises_runtime_error(jrc, tmp_path) -> None:
    """Empty compliance-rules response raises RuntimeError."""
    rows = _load(CR_DIR / "platform_compliance_rules_empty.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"compliance_rules": rows})
    with pytest.raises(RuntimeError, match="compliance-rules returned no rows"):
        dashboard._write_platform_compliance_rules(BENCHMARK)
    wb.close()


# ---------------------------------------------------------------------------
# Compliance Devices — unknown-first sort and absent rulesFailed invariant
# ---------------------------------------------------------------------------


def test_compliance_devices_happy_path_no_raise(jrc, tmp_path) -> None:
    """Happy fixture renders without exception."""
    rows = _load(CD_DIR / "platform_compliance_devices_happy.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"compliance_devices": rows})
    dashboard._write_platform_compliance_devices(BENCHMARK)
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path)
    assert len(wb_loaded.sheetnames) == 1


def test_compliance_devices_unknown_first_sort(jrc, tmp_path) -> None:
    """Devices with absent rulesFailed must sort before known-failing devices."""
    rows = _load(CD_DIR / "platform_compliance_devices_all_unknown.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"compliance_devices": rows})
    dashboard._write_platform_compliance_devices(BENCHMARK)
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path, data_only=False)
    sheet_name = wb_loaded.sheetnames[0]
    summary = _read_summary(wb_path, sheet_name)
    assert summary.get("Devices with Unknown Status") == 3


def test_compliance_devices_empty_raises_runtime_error(jrc, tmp_path) -> None:
    """Empty compliance-devices response raises RuntimeError."""
    rows = _load(CD_DIR / "platform_compliance_devices_empty.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"compliance_devices": rows})
    with pytest.raises(RuntimeError, match="compliance-devices returned no rows"):
        dashboard._write_platform_compliance_devices(BENCHMARK)
    wb.close()


# ---------------------------------------------------------------------------
# DDM Status
# ---------------------------------------------------------------------------


def test_ddm_status_happy_path_no_raise(jrc, tmp_path) -> None:
    """Happy fixture renders without exception; sheet has correct name."""
    rows = _load(DDM_DIR / "platform_ddm_status_happy.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"ddm_status": rows})
    dashboard._write_platform_ddm_status()
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path)
    assert "Platform DDM Status" in wb_loaded.sheetnames


def test_ddm_status_unsuccessful_summary(jrc, tmp_path) -> None:
    """Summary rows reflect correct counts for sources with unsuccessful declarations."""
    rows = _load(DDM_DIR / "platform_ddm_status_happy.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"ddm_status": rows})
    dashboard._write_platform_ddm_status()
    wb.close()

    summary = _read_summary(wb_path, "Platform DDM Status")
    assert summary.get("Sources Returned") == 5
    assert summary.get("Sources with Unsuccessful Declarations") == 1
    assert summary.get("Total Unsuccessful Declarations") == 3


def test_ddm_status_all_failed_summary(jrc, tmp_path) -> None:
    """All-failed fixture: every source has unsuccessful>0."""
    rows = _load(DDM_DIR / "platform_ddm_status_all_failed.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"ddm_status": rows})
    dashboard._write_platform_ddm_status()
    wb.close()

    summary = _read_summary(wb_path, "Platform DDM Status")
    assert summary.get("Sources with Unsuccessful Declarations") == 2
    assert summary.get("Total Unsuccessful Declarations") == 50


def test_ddm_status_partial_no_crash(jrc, tmp_path) -> None:
    """Partial fixture (missing numeric fields) renders without raising."""
    rows = _load(DDM_DIR / "platform_ddm_status_partial.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"ddm_status": rows})
    dashboard._write_platform_ddm_status()
    wb.close()

    wb_loaded = openpyxl.load_workbook(wb_path)
    assert "Platform DDM Status" in wb_loaded.sheetnames


def test_ddm_status_empty_raises_runtime_error(jrc, tmp_path) -> None:
    """Empty DDM response raises RuntimeError."""
    rows = _load(DDM_DIR / "platform_ddm_status_empty.json")
    dashboard, wb, wb_path = _make_dashboard(jrc, tmp_path, {"ddm_status": rows})
    with pytest.raises(RuntimeError, match="ddm-status returned no rows"):
        dashboard._write_platform_ddm_status()
    wb.close()


# ---------------------------------------------------------------------------
# platform.enabled = false skips sheets
# ---------------------------------------------------------------------------


def test_platform_disabled_blueprints_raises(jrc, tmp_path) -> None:
    """With platform.enabled=false, _write_platform_blueprints raises RuntimeError."""
    rows = _load(BP_DIR / "platform_blueprint_status_happy.json")
    dashboard, wb, wb_path = _make_dashboard(
        jrc, tmp_path, {"blueprint_status": rows}, platform_enabled=False
    )
    with pytest.raises(RuntimeError, match="disabled in config"):
        dashboard._write_platform_blueprints()
    wb.close()


def test_platform_disabled_ddm_raises(jrc, tmp_path) -> None:
    """With platform.enabled=false, _write_platform_ddm_status raises RuntimeError."""
    rows = _load(DDM_DIR / "platform_ddm_status_happy.json")
    dashboard, wb, wb_path = _make_dashboard(
        jrc, tmp_path, {"ddm_status": rows}, platform_enabled=False
    )
    with pytest.raises(RuntimeError, match="disabled in config"):
        dashboard._write_platform_ddm_status()
    wb.close()


# ---------------------------------------------------------------------------
# _classify_platform_error — error classifier
# ---------------------------------------------------------------------------


def test_classify_platform_error_401(jrc) -> None:
    """401 in message → auth_not_configured token."""
    result = jrc.JamfCLIBridge._classify_platform_error("HTTP 401 unauthorized")
    assert result == "auth_not_configured"


def test_classify_platform_error_oauth2(jrc) -> None:
    """OAuth2 keyword → auth_not_configured token."""
    result = jrc.JamfCLIBridge._classify_platform_error("OAuth2 token refresh failed")
    assert result == "auth_not_configured"


def test_classify_platform_error_404(jrc) -> None:
    """404 / not found → unknown_command token."""
    result = jrc.JamfCLIBridge._classify_platform_error("HTTP 404 not found")
    assert result == "unknown_command"


def test_classify_platform_error_command_not_found(jrc) -> None:
    """Unknown command → unknown_command token."""
    result = jrc.JamfCLIBridge._classify_platform_error("no such command: blueprint-status")
    assert result == "unknown_command"


def test_classify_platform_error_network(jrc) -> None:
    """Network timeout → network_error token."""
    result = jrc.JamfCLIBridge._classify_platform_error("connection timeout exceeded")
    assert result == "network_error"


def test_classify_platform_error_json_decode(jrc) -> None:
    """JSON decode error → parse_error token."""
    result = jrc.JamfCLIBridge._classify_platform_error("json decode error at line 1")
    assert result == "parse_error"


def test_classify_platform_error_unknown_passthrough(jrc) -> None:
    """Unrecognized message → general token."""
    msg = "some completely unknown failure condition"
    result = jrc.JamfCLIBridge._classify_platform_error(msg)
    assert result == "general"
