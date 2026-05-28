"""Tests for the v2.1.0 PR-3 mobile-details opt-out and supervision sheet.

Covers:
- `_normalized_skip_types` normalises hyphen/underscore spellings and lower-cases.
- `_collect_jamf_cli_commands` drops the Mobile Inventory fetch when
  ``jamf_cli.collect_skip`` contains ``mobile-details`` (either spelling).
- `_normalize_mobile_inventory_row` surfaces an "Enrollment Method" label and a
  "Managed Apps" count derived from `general.deviceOwnershipType`,
  `general.enrollmentMethodPrestage.profileName`, and the top-level
  `applications` array.
- `CoreDashboard._write_mobile_supervision_status` renders per-family
  supervision counts; raises ``RuntimeError`` (silently `[skip]`-ed by
  ``write_all``) when no rich mobile data exists.
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest
import xlsxwriter
import yaml


class _BridgeStub:
    """Fake JamfCLIBridge — every method returns ``None`` for `_collect_jamf_cli_commands`."""

    def __getattr__(self, _name: str):
        return lambda *args, **kwargs: None


def _config_with_skip(jrc, tmp_path: Path, skip_values: list[str]):
    cfg_path = tmp_path / "config.yaml"
    cfg_path.write_text(yaml.safe_dump({"jamf_cli": {"collect_skip": skip_values}}))
    return jrc.Config(str(cfg_path))


def _build_dashboard(jrc, tmp_path: Path, mobile_items: list[dict]):
    config = jrc.Config(jrc.Config._WORKSPACE_INIT_DEFAULTS_NAME)
    bridge = MagicMock()
    bridge.mobile_device_inventory_details.return_value = mobile_items
    bridge.mobile_devices_list.return_value = []
    wb_path = str(tmp_path / "mobile-details.xlsx")
    wb = xlsxwriter.Workbook(wb_path, {"remove_timezone": True})
    fmts = jrc._build_formats(wb)
    dashboard = jrc.CoreDashboard(config, bridge, wb, fmts)
    return dashboard, wb, wb_path


# ---------------------------------------------------------------------------
# Helper: skip-types normalisation
# ---------------------------------------------------------------------------


def test_normalized_skip_types_lowercases_and_dashes(jrc, tmp_path) -> None:
    config = _config_with_skip(
        jrc, tmp_path, ["Mobile_Details", "DEVICE-SECURITY-STATE", "update_status"]
    )
    skip = jrc._normalized_skip_types(config)
    assert "mobile-details" in skip
    assert "device-security-state" in skip
    assert "update-status" in skip


def test_normalized_skip_types_empty_when_missing(jrc, tmp_path) -> None:
    config = jrc.Config(jrc.Config._WORKSPACE_INIT_DEFAULTS_NAME)
    assert jrc._normalized_skip_types(config) == set()


# ---------------------------------------------------------------------------
# collect_skip drops the Mobile Inventory call entirely
# ---------------------------------------------------------------------------


def _collected_labels(jrc, config) -> list[str]:
    bridge = _BridgeStub()
    commands = jrc._collect_jamf_cli_commands(config, bridge, True)
    return [label for label, _ in commands]


def test_collect_skip_mobile_details_drops_mobile_inventory(jrc, tmp_path) -> None:
    config = _config_with_skip(jrc, tmp_path, ["mobile-details"])
    labels = _collected_labels(jrc, config)
    assert "Mobile Inventory" not in labels
    # Sibling mobile commands stay — only the heavy details fetch is gated.
    assert "Mobile Device List" in labels
    assert "Mobile Config Profiles" in labels


def test_collect_skip_mobile_details_normalises_underscore(jrc, tmp_path) -> None:
    config = _config_with_skip(jrc, tmp_path, ["mobile_details"])
    assert "Mobile Inventory" not in _collected_labels(jrc, config)


def test_collect_skip_without_mobile_details_keeps_call(jrc, tmp_path) -> None:
    config = _config_with_skip(jrc, tmp_path, [])
    assert "Mobile Inventory" in _collected_labels(jrc, config)


# ---------------------------------------------------------------------------
# Enrollment-method + managed-apps row normalisation
# ---------------------------------------------------------------------------


def test_enrollment_label_maps_known_enums(jrc) -> None:
    assert jrc._mobile_enrollment_label("Institutional", None) == "ADE / Institutional"
    assert jrc._mobile_enrollment_label("UserEnrollment", None) == "User Enrollment"
    assert (
        jrc._mobile_enrollment_label("AccountDrivenUserEnrollment", None)
        == "Account-Driven User Enrollment"
    )
    assert (
        jrc._mobile_enrollment_label("AccountDrivenDeviceEnrollment", None)
        == "Account-Driven Device Enrollment"
    )


def test_enrollment_label_appends_prestage_when_present(jrc) -> None:
    label = jrc._mobile_enrollment_label("Institutional", "All Mobiles")
    assert label == "ADE / Institutional (Prestage: All Mobiles)"


def test_enrollment_label_falls_back_to_raw_unknown_value(jrc) -> None:
    assert jrc._mobile_enrollment_label("NewServerSideEnum", None) == "NewServerSideEnum"


def test_enrollment_label_empty_for_empty_input(jrc) -> None:
    assert jrc._mobile_enrollment_label("", None) == ""
    assert jrc._mobile_enrollment_label(None, "Ignored") == ""


def test_normalize_row_surfaces_enrollment_and_managed_apps(jrc) -> None:
    item = {
        "mobileDeviceId": "1",
        "deviceType": "iOS",
        "general": {
            "displayName": "iPad-Lab-001",
            "serialNumber": "DMPH12345678",
            "osVersion": "18.2.1",
            "managed": True,
            "supervised": True,
            "deviceOwnershipType": "Institutional",
            "enrollmentMethodPrestage": {
                "mobileDevicePrestageId": "5",
                "profileName": "Education Prestage",
            },
        },
        "applications": [
            {"identifier": "com.apple.calculator", "name": "Calculator"},
            {"identifier": "com.apple.maps", "name": "Maps"},
            {"identifier": "com.example.app", "name": "Example"},
        ],
    }
    row = jrc._normalize_mobile_inventory_row(item)
    assert row["Enrollment Method"] == "ADE / Institutional (Prestage: Education Prestage)"
    assert row["Managed Apps"] == 3
    assert row["Ownership"] == "Institutional"


def test_normalize_row_handles_null_applications(jrc) -> None:
    item = {
        "general": {
            "displayName": "iPhone-001",
            "serialNumber": "ABCD",
            "deviceOwnershipType": "UserEnrollment",
        },
        "applications": None,
    }
    row = jrc._normalize_mobile_inventory_row(item)
    assert row["Managed Apps"] == 0
    assert row["Enrollment Method"] == "User Enrollment"


# ---------------------------------------------------------------------------
# Mobile Supervision Status sheet
# ---------------------------------------------------------------------------


_SUPERVISED_FIXTURE: list[dict] = [
    {
        "mobileDeviceId": "1",
        "deviceType": "iPad",
        "general": {
            "displayName": "iPad-Edu-1",
            "serialNumber": "IPAD0001",
            "osVersion": "18.2.1",
            "model": "iPad 10th",
            "managed": True,
            "supervised": True,
            "deviceOwnershipType": "Institutional",
        },
    },
    {
        "mobileDeviceId": "2",
        "deviceType": "iPad",
        "general": {
            "displayName": "iPad-Edu-2",
            "serialNumber": "IPAD0002",
            "osVersion": "18.2.1",
            "model": "iPad 10th",
            "managed": True,
            "supervised": False,
            "deviceOwnershipType": "Institutional",
        },
    },
    {
        "mobileDeviceId": "3",
        "deviceType": "iPhone",
        "general": {
            "displayName": "iPhone-Exec-1",
            "serialNumber": "IPHN0001",
            "osVersion": "18.1.1",
            "model": "iPhone 15",
            "managed": True,
            "supervised": True,
            "deviceOwnershipType": "UserEnrollment",
        },
    },
]


def test_write_mobile_supervision_status_renders_per_family_rows(jrc, tmp_path) -> None:
    dashboard, wb, wb_path = _build_dashboard(jrc, tmp_path, _SUPERVISED_FIXTURE)
    dashboard._write_mobile_supervision_status()
    wb.close()

    book = openpyxl.load_workbook(wb_path, data_only=False)
    ws = book["Mobile Supervision Status"]
    headers = [ws.cell(row=4, column=col).value for col in range(1, 6)]
    assert headers == ["Device Family", "Total", "Supervised", "Unsupervised", "% Supervised"]

    rows = {ws.cell(row=r, column=1).value: r for r in range(5, ws.max_row + 1)}
    assert "iPad" in rows
    assert "iPhone" in rows

    ipad_row = rows["iPad"]
    assert ws.cell(row=ipad_row, column=2).value == 2  # total
    assert ws.cell(row=ipad_row, column=3).value == 1  # supervised
    assert ws.cell(row=ipad_row, column=4).value == 1  # unsupervised
    assert ws.cell(row=ipad_row, column=5).value == pytest.approx(0.5)

    iphone_row = rows["iPhone"]
    assert ws.cell(row=iphone_row, column=2).value == 1
    assert ws.cell(row=iphone_row, column=3).value == 1
    assert ws.cell(row=iphone_row, column=5).value == pytest.approx(1.0)


def test_write_mobile_supervision_status_raises_when_no_rows(jrc, tmp_path) -> None:
    dashboard, wb, _ = _build_dashboard(jrc, tmp_path, [])
    with pytest.raises(RuntimeError):
        dashboard._write_mobile_supervision_status()
    wb.close()


# ---------------------------------------------------------------------------
# Mobile Inventory sheet picks up the new columns
# ---------------------------------------------------------------------------


def test_mobile_inventory_sheet_includes_new_columns(jrc, tmp_path) -> None:
    dashboard, wb, wb_path = _build_dashboard(jrc, tmp_path, _SUPERVISED_FIXTURE)
    dashboard._write_mobile_inventory()
    wb.close()

    book = openpyxl.load_workbook(wb_path, data_only=False)
    ws = book["Mobile Inventory"]

    # Header row sits below the title block + summary stats — find by scanning.
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Jamf Pro ID":
            header_row = r
            break
    assert header_row is not None, "Mobile Inventory header row not found"

    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Enrollment Method" in headers
    assert "Managed Apps" in headers
