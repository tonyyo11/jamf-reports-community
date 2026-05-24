"""Tests for the v2.1.0 device security state opt-out and Excel sheet.

Covers:
- `_inventory_computer_sections_without_security` returns sections minus SECURITY.
- `_collect_jamf_cli_commands` routes the Computer Inventory call to the
  sectionless variant when `jamf_cli.collect_skip` contains
  ``device-security-state`` (underscore or hyphen).
- `CoreDashboard._write_device_security_state` renders one row per device with
  red/green/neutral cell shading from cached `computers-list` JSON.
- The same writer raises ``RuntimeError`` (silently skipped by ``write_all``)
  when no record carries any security value — e.g. on tenants that opted out
  of the SECURITY inventory section.
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest
import xlsxwriter
import yaml


class _BridgeStub:
    """Fake JamfCLIBridge — every method returns a no-op lambda."""

    def __getattr__(self, _name: str):
        return lambda *args, **kwargs: None


def _config_with_skip(jrc, tmp_path: Path, skip_values: list[str]):
    cfg_path = tmp_path / "config.yaml"
    cfg_path.write_text(yaml.safe_dump({"jamf_cli": {"collect_skip": skip_values}}))
    return jrc.Config(str(cfg_path))


def _build_dashboard(jrc, tmp_path: Path, computers: list[dict]):
    config = jrc.Config(jrc.Config._WORKSPACE_INIT_DEFAULTS_NAME)
    bridge = MagicMock()
    bridge.computers_list.return_value = computers
    wb_path = str(tmp_path / "device-security.xlsx")
    wb = xlsxwriter.Workbook(wb_path, {"remove_timezone": True})
    fmts = jrc._build_formats(wb)
    dashboard = jrc.CoreDashboard(config, bridge, wb, fmts)
    return dashboard, wb, wb_path


# ---------------------------------------------------------------------------
# Helper: SECURITY-section omission
# ---------------------------------------------------------------------------


def test_inventory_sections_without_security_drops_security(jrc) -> None:
    full = jrc._inventory_computer_sections()
    pared = jrc._inventory_computer_sections_without_security()
    assert "SECURITY" in full
    assert "SECURITY" not in pared
    assert set(pared) == set(full) - {"SECURITY"}


# ---------------------------------------------------------------------------
# collect_skip recognises device-security-state
# ---------------------------------------------------------------------------


def _captured_sections(jrc, config) -> list[str]:
    bridge = _BridgeStub()
    captured: dict[str, list[str]] = {}

    def _capture(sections):
        captured["sections"] = list(sections)
        return None

    bridge.computers_list = _capture  # type: ignore[method-assign]
    commands = jrc._collect_jamf_cli_commands(config, bridge, True)
    inventory = next(cmd for label, cmd in commands if label == "Computer Inventory")
    inventory()
    return captured["sections"]


def test_collect_skip_device_security_state_drops_security_section(
    jrc, tmp_path
) -> None:
    config = _config_with_skip(jrc, tmp_path, ["device-security-state"])
    sections = _captured_sections(jrc, config)
    assert "SECURITY" not in sections
    assert "GENERAL" in sections


def test_collect_skip_device_security_state_normalizes_underscore(
    jrc, tmp_path
) -> None:
    config = _config_with_skip(jrc, tmp_path, ["device_security_state"])
    sections = _captured_sections(jrc, config)
    assert "SECURITY" not in sections


def test_collect_skip_without_device_security_state_keeps_section(
    jrc, tmp_path
) -> None:
    config = _config_with_skip(jrc, tmp_path, [])
    sections = _captured_sections(jrc, config)
    assert "SECURITY" in sections


# ---------------------------------------------------------------------------
# Sheet rendering — happy path
# ---------------------------------------------------------------------------


_HAPPY_COMPUTERS: list[dict] = [
    {
        "general": {"id": "1", "name": "Mac-OK"},
        "hardware": {"serialNumber": "OK0001"},
        "diskEncryption": {
            "fileVault2Enabled": True,
            "bootPartitionEncryptionDetails": {"partitionFileVault2State": "ENCRYPTED"},
        },
        "security": {
            "sipStatus": "ENABLED",
            "firewallEnabled": True,
            "gatekeeperStatus": "APP_STORE_AND_IDENTIFIED_DEVELOPERS",
            "bootstrapTokenEscrowed": True,
        },
    },
    {
        "general": {"id": "2", "name": "Mac-Bad"},
        "hardware": {"serialNumber": "BAD0002"},
        "diskEncryption": {
            "fileVault2Enabled": False,
            "bootPartitionEncryptionDetails": {"partitionFileVault2State": "UNENCRYPTED"},
        },
        "security": {
            "sipStatus": "DISABLED",
            "firewallEnabled": False,
            "gatekeeperStatus": "DISABLED",
            "bootstrapTokenEscrowed": False,
        },
    },
    {
        # Missing SECURITY entirely — excluded from the sheet.
        "general": {"id": "3", "name": "Mac-NoData"},
        "hardware": {"serialNumber": "NODATA0003"},
    },
]


def test_write_device_security_state_renders_rows(jrc, tmp_path) -> None:
    dashboard, wb, wb_path = _build_dashboard(jrc, tmp_path, _HAPPY_COMPUTERS)
    dashboard._write_device_security_state()
    wb.close()

    book = openpyxl.load_workbook(wb_path, data_only=False)
    ws = book["Device Security State"]
    headers = [ws.cell(row=4, column=col).value for col in range(1, 8)]
    assert headers == [
        "Device Name", "Serial", "FileVault", "SIP",
        "Firewall", "Gatekeeper", "Bootstrap Token",
    ]

    names = [ws.cell(row=r, column=1).value for r in range(5, ws.max_row + 1)]
    assert names == ["Mac-Bad", "Mac-OK"]  # alphabetised, no-security row absent

    by_name = {ws.cell(row=r, column=1).value: r for r in range(5, ws.max_row + 1)}
    bad_row = by_name["Mac-Bad"]
    ok_row = by_name["Mac-OK"]

    assert ws.cell(row=bad_row, column=3).value.upper() in {"FALSE", "UNENCRYPTED"}
    assert ws.cell(row=bad_row, column=4).value == "DISABLED"
    assert ws.cell(row=ok_row, column=3).value.upper() in {"TRUE", "ENCRYPTED"}
    assert ws.cell(row=ok_row, column=4).value == "ENABLED"


def test_write_device_security_state_raises_when_no_security_data(
    jrc, tmp_path
) -> None:
    no_security = [
        {
            "general": {"id": "1", "name": "Mac-Empty"},
            "hardware": {"serialNumber": "EMPTY"},
        }
    ]
    dashboard, wb, _ = _build_dashboard(jrc, tmp_path, no_security)
    with pytest.raises(RuntimeError, match="no device security state"):
        dashboard._write_device_security_state()
    wb.close()
