"""Tests for SOFAFeedClient and the OS Currency sheet."""

from __future__ import annotations

import subprocess
from datetime import date

import openpyxl


def _sofa_config(config_factory, fixtures_root, *, enabled=True, max_age=10_000):
    """Return a Config whose SOFA cache points at the committed fixtures."""
    config = config_factory("dummy.yaml")
    config._data["sofa"] = {
        "enabled": enabled,
        "cache_dir": str(fixtures_root / "sofa"),
        "max_cache_age_hours": max_age,
        "timeout_seconds": 30,
        "platforms": ["macos", "ios", "tvos", "watchos"],
    }
    return config


def _empty_cache_config(config_factory, tmp_path):
    """Return a Config whose SOFA cache dir is empty (no fixtures)."""
    config = config_factory("dummy.yaml")
    config._data["sofa"] = {
        "enabled": True,
        "cache_dir": str(tmp_path / "empty-cache"),
        "max_cache_age_hours": 0,
        "timeout_seconds": 30,
        "platforms": ["macos"],
    }
    return config


def test_parses_all_four_platform_fixtures(jrc, config_factory, fixtures_root) -> None:
    config = _sofa_config(config_factory, fixtures_root)
    client = jrc.SOFAFeedClient(config)
    for platform in ("macos", "ios", "tvos", "watchos"):
        feed = client.fetch(platform)
        assert feed is not None, platform
        assert "OSVersions" in feed
        assert feed["OSVersions"], platform


def test_date_parser_handles_iso_timestamp_and_date_only(jrc) -> None:
    assert jrc._sofa_parse_date("2026-06-01T00:00:00Z") == date(2026, 6, 1)
    assert jrc._sofa_parse_date("2026-05-11") == date(2026, 5, 11)
    assert jrc._sofa_parse_date("") is None
    assert jrc._sofa_parse_date(None) is None
    assert jrc._sofa_parse_date("not-a-date") is None


def test_latest_versions_returns_correct_product_versions(jrc, config_factory, fixtures_root) -> None:
    config = _sofa_config(config_factory, fixtures_root)
    client = jrc.SOFAFeedClient(config, reference_date=date(2026, 6, 2))
    rows = client.latest_versions()

    by_platform = {}
    for row in rows:
        by_platform.setdefault(row["platform"], row)

    assert by_platform["macOS"]["product_version"] == "26.5.1"
    # Released 2026-06-01, reference 2026-06-02 -> 1 day.
    assert by_platform["macOS"]["days_since_release"] == 1
    assert by_platform["iOS / iPadOS"]["product_version"] == "26.5"
    assert by_platform["tvOS"]["product_version"] == "26.5"
    assert by_platform["watchOS"]["product_version"] == "26.5"


def test_days_since_release_uses_reference_date(jrc, config_factory, fixtures_root) -> None:
    config = _sofa_config(config_factory, fixtures_root)
    client = jrc.SOFAFeedClient(config, reference_date=date(2026, 6, 11))
    macos = next(r for r in client.latest_versions() if r["platform"] == "macOS")
    # 2026-06-01 -> 2026-06-11 is 10 days.
    assert macos["days_since_release"] == 10


def test_disabled_returns_empty_and_does_not_fetch(
    jrc, config_factory, fixtures_root, monkeypatch
) -> None:
    config = _sofa_config(config_factory, fixtures_root, enabled=False)
    client = jrc.SOFAFeedClient(config)

    def _boom(*args, **kwargs):
        raise AssertionError("curl must not run when SOFA is disabled")

    monkeypatch.setattr(subprocess, "run", _boom)

    assert client.enabled is False
    assert client.latest_versions() == []
    assert client.fetch("macos") is None


def test_fetch_failure_without_cache_returns_none(
    jrc, config_factory, tmp_path, monkeypatch, real_sofa_curl
) -> None:
    config = _empty_cache_config(config_factory, tmp_path)
    client = jrc.SOFAFeedClient(config)
    monkeypatch.setattr(jrc.SOFAFeedClient, "_curl_fetch", real_sofa_curl)

    def _fail(*args, **kwargs):
        raise subprocess.CalledProcessError(22, "curl")

    monkeypatch.setattr(subprocess, "run", _fail)
    assert client.fetch("macos") is None


def test_fetch_failure_with_cache_uses_cache(
    jrc, config_factory, fixtures_root, monkeypatch, real_sofa_curl
) -> None:
    # max_age 0 forces a live attempt every time; the live call fails, so the
    # committed fixture is used as the cache fallback.
    config = _sofa_config(config_factory, fixtures_root, max_age=0)
    client = jrc.SOFAFeedClient(config)
    monkeypatch.setattr(jrc.SOFAFeedClient, "_curl_fetch", real_sofa_curl)

    def _fail(*args, **kwargs):
        raise subprocess.CalledProcessError(22, "curl")

    monkeypatch.setattr(subprocess, "run", _fail)
    feed = client.fetch("macos")
    assert feed is not None
    assert feed["OSVersions"][0]["Latest"]["ProductVersion"] == "26.5.1"


def test_curl_missing_returns_none_gracefully(
    jrc, config_factory, tmp_path, monkeypatch, real_sofa_curl
) -> None:
    config = _empty_cache_config(config_factory, tmp_path)
    client = jrc.SOFAFeedClient(config)
    monkeypatch.setattr(jrc.SOFAFeedClient, "_curl_fetch", real_sofa_curl)

    def _missing(*args, **kwargs):
        raise FileNotFoundError("curl")

    monkeypatch.setattr(subprocess, "run", _missing)
    assert client.fetch("macos") is None


class _CoreBridge:
    """Minimal bridge double for OS Currency sheet rendering."""

    def __init__(self, security=None, mobile=None):
        self._security = security or []
        self._mobile = mobile or []

    def security_report(self):
        return self._security

    def mobile_device_inventory_details(self):
        return self._mobile

    def mobile_devices_list(self):
        return self._mobile


def _write_one_sheet(jrc, config, bridge, writer_name, tmp_path):
    """Build a workbook, run a single CoreDashboard writer, return the path."""
    import xlsxwriter

    out = tmp_path / "os-currency.xlsx"
    wb = xlsxwriter.Workbook(str(out))
    fmts = jrc._build_formats(wb)
    dash = jrc.CoreDashboard(config, bridge, wb, fmts)
    getattr(dash, writer_name)()
    wb.close()
    return out


def test_os_currency_sheet_renders_with_sofa_data(
    jrc, config_factory, fixtures_root, tmp_path
) -> None:
    config = _sofa_config(config_factory, fixtures_root)
    security = [
        {"section": "summary", "data": {"total_devices": 3}},
        {"section": "os_version", "os_version": "26.5.1", "count": 2},
        {"section": "os_version", "os_version": "15.7.0", "count": 1},
    ]
    bridge = _CoreBridge(security=security)
    path = _write_one_sheet(jrc, config, bridge, "_write_os_currency", tmp_path)

    wb = openpyxl.load_workbook(path)
    sheet = wb["OS Currency"]
    values = [cell.value for row in sheet.iter_rows() for cell in row]
    assert "Latest Version" in values
    assert "26.5.1" in values
    assert "Fleet On Latest" in values


def test_os_currency_sheet_renders_note_without_sofa_data(
    jrc, config_factory, fixtures_root, tmp_path
) -> None:
    config = _sofa_config(config_factory, fixtures_root, enabled=False)
    bridge = _CoreBridge()
    path = _write_one_sheet(jrc, config, bridge, "_write_os_currency", tmp_path)

    wb = openpyxl.load_workbook(path)
    assert "OS Currency" in wb.sheetnames
    sheet = wb["OS Currency"]
    values = [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value]
    assert any("SOFA feed unavailable" in v for v in values)


def test_os_currency_sheet_shows_eol_devices_row(
    jrc, config_factory, fixtures_root, tmp_path
) -> None:
    """Devices on majors older than every SOFA family appear as an EOL row.

    Regression: a fleet running only EOL versions (e.g. Jamf's demo tenant on
    10.9.5) previously rendered all zeros, making the sheet look broken and
    hiding the most security-relevant finding.
    """
    config = _sofa_config(config_factory, fixtures_root)
    security = [
        {"section": "summary", "data": {"total_devices": 101}},
        {"section": "os_version", "os_version": "10.9.5", "count": 100},
        {"section": "os_version", "os_version": "11.6.2", "count": 1},
    ]
    bridge = _CoreBridge(security=security)
    path = _write_one_sheet(jrc, config, bridge, "_write_os_currency", tmp_path)

    wb = openpyxl.load_workbook(path)
    sheet = wb["OS Currency"]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
    eol_rows = [r for r in rows if r and r[1] == "Out of support (EOL)"]
    assert len(eol_rows) == 1, "expected exactly one macOS EOL row"
    eol = eol_rows[0]
    assert eol[0] == "macOS"
    # Fleet Behind column carries the EOL device count (101 = 100 + 1)
    assert eol[8] == 101


def test_fleet_eol_count_ignores_supported_majors(jrc) -> None:
    counts = {"15.7.7": 5, "26.5.1": 3, "10.9.5": 7, "11.6.2": 2}
    family_majors = {26, 15, 14, 13, 12}
    devices, versions = jrc.CoreDashboard._fleet_eol_count(family_majors, counts)
    assert devices == 9          # 7 + 2 below major 12
    assert versions == 2
    # No families -> no EOL classification possible
    assert jrc.CoreDashboard._fleet_eol_count(set(), counts) == (0, 0)
