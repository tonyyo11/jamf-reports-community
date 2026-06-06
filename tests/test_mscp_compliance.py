"""Tests for mSCP/STIG compliance banding parity with the Swift engine.

Validates the per-baseline band math, real compliancePct, summary mscpBands
wiring, sheet skip-when-no-data behavior, and config key parity. Band semantics
must stay byte-identical to MSCPComplianceService + ComplianceBandingService in
the Swift app.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Optional

import pytest


def _config(jrc, tmp_path: Path, compliance: dict[str, Any]):
    """Build a Config from a minimal YAML file containing a compliance block."""
    import yaml

    path = tmp_path / "config.yaml"
    path.write_text(yaml.safe_dump({"compliance": compliance}), encoding="utf-8")
    return jrc.Config(str(path))


def _row(value: Any, ea_name: str = "Failed Count", device: str = "mac-1", **extra):
    row = {"ea_name": ea_name, "value": value, "device": device}
    row.update(extra)
    return row


class _StubBridge:
    """Minimal JamfCLIBridge stand-in for summary-wiring tests."""

    def __init__(self, rows: Optional[list], available: bool = True):
        self._rows = rows
        self._available = available
        self.calls = 0

    def is_available(self) -> bool:
        return self._available

    def ea_results_report(self, include_all: bool = True):
        self.calls += 1
        return self._rows


# ---------------------------------------------------------------------------
# Value parser parity (Swift AnyCodable.intValue + >= 0 guard)
# ---------------------------------------------------------------------------


def test_int_value_parses_int_float_and_string(jrc):
    assert jrc._mscp_int_value(5) == 5
    assert jrc._mscp_int_value(0) == 0
    assert jrc._mscp_int_value(3.9) == 3  # truncate toward zero, matches Int(Double)
    assert jrc._mscp_int_value("12") == 12


def test_int_value_rejects_swift_incompatible_strings(jrc):
    # Swift Int("5.0") == nil; our parser must agree -> No Data, not Low.
    assert jrc._mscp_int_value("5.0") is None
    assert jrc._mscp_int_value("1e3") is None
    assert jrc._mscp_int_value(" 5 ") is None
    assert jrc._mscp_int_value("abc") is None
    assert jrc._mscp_int_value("") is None


def test_int_value_rejects_bool_and_negative(jrc):
    # bool is an int subclass in Python; Swift intValue is nil for JSON Bool.
    assert jrc._mscp_int_value(True) is None
    assert jrc._mscp_int_value(False) is None
    assert jrc._mscp_int_value(-1) is None
    assert jrc._mscp_int_value("-4") is None
    assert jrc._mscp_int_value(None) is None


# ---------------------------------------------------------------------------
# Band classification boundaries
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "failures,expected",
    [
        (None, "noData"),
        (-1, "noData"),
        (0, "pass"),
        (1, "low"),
        (10, "low"),
        (11, "medLow"),
        (30, "medLow"),
        (31, "medium"),
        (50, "medium"),
        (51, "high"),
        (9999, "high"),
    ],
)
def test_band_key_boundaries(jrc, failures, expected):
    assert jrc._mscp_band_key(failures) == expected


# ---------------------------------------------------------------------------
# resolvedBaselines synthesis
# ---------------------------------------------------------------------------


def test_resolved_baselines_uses_explicit_list(jrc):
    cfg = {
        "baselines": [
            {"name": "NIST", "failures_count_column": "EA-NIST"},
            {"name": "STIG", "failures_count_column": "EA-STIG"},
        ]
    }
    resolved = jrc._mscp_resolved_baselines(cfg)
    assert [b["name"] for b in resolved] == ["NIST", "STIG"]
    assert resolved[0]["failures_count_column"] == "EA-NIST"


def test_resolved_baselines_synthesizes_from_legacy_column(jrc):
    cfg = {"failures_count_column": "Failed Count", "baseline_label": "mSCP"}
    resolved = jrc._mscp_resolved_baselines(cfg)
    assert resolved == [{"name": "mSCP", "failures_count_column": "Failed Count"}]


def test_resolved_baselines_empty_when_unconfigured(jrc):
    assert jrc._mscp_resolved_baselines({}) == []
    assert jrc._mscp_resolved_baselines({"baselines": []}) == []


def test_resolved_baselines_default_name_when_label_blank(jrc):
    cfg = {"failures_count_column": "Failed Count", "baseline_label": ""}
    assert jrc._mscp_resolved_baselines(cfg)[0]["name"] == "Compliance"


# ---------------------------------------------------------------------------
# Band distribution + compliancePct math
# ---------------------------------------------------------------------------


def test_evaluate_baseline_band_distribution_and_pct(jrc):
    rows = [
        _row(0, device="d1"),     # pass
        _row(5, device="d2"),     # low
        _row(20, device="d3"),    # medLow
        _row(40, device="d4"),    # medium
        _row(99, device="d5"),    # high
        _row(0, device="d6"),     # pass
        _row("bad", device="d7"), # unparseable -> noData (in universe)
        _row(0, device="d8", ea_name="Other EA"),  # other EA -> noData for this baseline
    ]
    res = jrc._mscp_evaluate_baseline(rows, "Failed Count")
    assert res["total"] == 8  # universe = 8 distinct devices
    assert res["bands"] == {
        "pass": 2,
        "low": 1,
        "medLow": 1,
        "medium": 1,
        "high": 1,
        "noData": 2,
    }
    assert res["noData"] == 2
    assert res["devices_with_data"] == 6
    assert res["pass"] == 2
    # compliancePct = pass / devicesWithData = 2 / 6
    assert res["compliance_pct"] == pytest.approx(2 / 6 * 100.0)


def test_evaluate_baseline_pct_none_when_no_devices_with_data(jrc):
    # Devices exist in the universe (other EA) but none for this baseline's EA.
    rows = [_row(0, device="d1", ea_name="Other"), _row(3, device="d2", ea_name="Other")]
    res = jrc._mscp_evaluate_baseline(rows, "Failed Count")
    assert res["total"] == 2
    assert res["devices_with_data"] == 0
    assert res["compliance_pct"] is None


def test_evaluate_baseline_identity_chain_and_case_insensitive_ea(jrc):
    rows = [
        {"computer_id": "C1", "ea_name": "failed count", "value": 0},   # ci EA match, id wins
        {"serial": "S2", "ea_name": "Failed Count", "value": 7},
        {"computer_name": "Name3", "ea_name": "Failed Count", "value": 0},
    ]
    res = jrc._mscp_evaluate_baseline(rows, "Failed Count")
    assert res["total"] == 3
    assert res["bands"]["pass"] == 2
    assert res["bands"]["low"] == 1


def test_evaluate_baseline_last_write_wins_per_device(jrc):
    # Same device with two parseable rows: later one wins (matches Swift dict assign).
    rows = [_row(0, device="dup"), _row(40, device="dup")]
    res = jrc._mscp_evaluate_baseline(rows, "Failed Count")
    assert res["total"] == 1
    assert res["bands"]["medium"] == 1
    assert res["bands"]["pass"] == 0


def test_evaluate_baseline_skips_rows_without_identity(jrc):
    rows = [_row(0, device=""), {"ea_name": "Failed Count", "value": 5}]
    res = jrc._mscp_evaluate_baseline(rows, "Failed Count")
    assert res["total"] == 0


# ---------------------------------------------------------------------------
# Summary mscpBands wiring
# ---------------------------------------------------------------------------


def test_apply_mscp_bands_sets_real_pct_and_proxy_false(jrc, tmp_path):
    cfg = _config(
        jrc, tmp_path, {"failures_count_column": "Failed Count", "baseline_label": "mSCP"}
    )
    rows = [_row(0, device="d1"), _row(0, device="d2"), _row(5, device="d3")]
    bridge = _StubBridge(rows)
    summary: dict[str, Any] = {"date": "2026-06-05", "compliancePct": 99.0}
    jrc._apply_mscp_bands_to_summary(summary, cfg, bridge)
    assert summary["complianceIsProxy"] is False
    # pass=2 / devicesWithData=3 = 66.7, overrides the prior CSV value of 99.0
    assert summary["compliancePct"] == pytest.approx(66.7)
    assert summary["mscpBands"]["mSCP"] == {
        "pass": 2,
        "low": 1,
        "medLow": 0,
        "medium": 0,
        "high": 0,
        "noData": 0,
        "total": 3,
    }


def test_apply_mscp_bands_keyed_per_baseline_name(jrc, tmp_path):
    cfg = _config(
        jrc,
        tmp_path,
        {
            "baselines": [
                {"name": "NIST", "failures_count_column": "EA-NIST"},
                {"name": "STIG", "failures_count_column": "EA-STIG"},
            ]
        },
    )
    rows = [
        _row(0, ea_name="EA-NIST", device="d1"),
        _row(60, ea_name="EA-STIG", device="d1"),
    ]
    bridge = _StubBridge(rows)
    summary: dict[str, Any] = {"date": "2026-06-05"}
    jrc._apply_mscp_bands_to_summary(summary, cfg, bridge)
    assert set(summary["mscpBands"]) == {"NIST", "STIG"}
    # Primary baseline (NIST) drives compliancePct: d1 pass, d1 noData for STIG.
    assert summary["mscpBands"]["NIST"]["pass"] == 1
    assert summary["mscpBands"]["STIG"]["high"] == 1
    assert summary["complianceIsProxy"] is False


def test_apply_mscp_bands_noop_when_bridge_unavailable(jrc, tmp_path):
    cfg = _config(jrc, tmp_path, {"failures_count_column": "Failed Count"})
    bridge = _StubBridge([_row(0)], available=False)
    summary: dict[str, Any] = {"date": "2026-06-05", "compliancePct": 50.0}
    jrc._apply_mscp_bands_to_summary(summary, cfg, bridge)
    assert "mscpBands" not in summary
    assert "complianceIsProxy" not in summary
    assert summary["compliancePct"] == 50.0


def test_apply_mscp_bands_noop_when_no_baseline_configured(jrc, tmp_path):
    cfg = _config(jrc, tmp_path, {"baselines": []})
    bridge = _StubBridge([_row(0)])
    summary: dict[str, Any] = {"date": "2026-06-05"}
    jrc._apply_mscp_bands_to_summary(summary, cfg, bridge)
    assert "mscpBands" not in summary


def test_apply_mscp_bands_pct_unchanged_when_no_devices_with_data(jrc, tmp_path):
    # Baseline configured, rows exist, but none match the EA -> keep proxy pct.
    cfg = _config(jrc, tmp_path, {"failures_count_column": "Failed Count"})
    bridge = _StubBridge([_row(0, ea_name="Other", device="d1")])
    summary: dict[str, Any] = {"date": "2026-06-05", "compliancePct": 88.0}
    jrc._apply_mscp_bands_to_summary(summary, cfg, bridge)
    # Bands still emitted (universe has a device), but pct/proxy untouched.
    assert "mscpBands" in summary
    assert summary["compliancePct"] == 88.0
    assert "complianceIsProxy" not in summary


# ---------------------------------------------------------------------------
# Dated snapshot loader (Compliance Trend source)
# ---------------------------------------------------------------------------


def test_load_ea_results_snapshots_sorted_by_date(jrc, tmp_path):
    results_dir = tmp_path / "ea-results"
    results_dir.mkdir()
    (results_dir / "ea-results_2026-06-03.json").write_text(
        json.dumps([_row(0, device="d1")]), encoding="utf-8"
    )
    (results_dir / "ea-results_2026-06-01.json").write_text(
        json.dumps([_row(5, device="d1")]), encoding="utf-8"
    )
    snaps = jrc._load_ea_results_snapshots(tmp_path)
    assert len(snaps) == 2
    assert snaps[0][0] < snaps[1][0]  # ascending
    assert snaps[0][1][0]["value"] == 5  # 2026-06-01 first


def test_load_ea_results_snapshots_empty_when_absent(jrc, tmp_path):
    assert jrc._load_ea_results_snapshots(tmp_path) == []


# ---------------------------------------------------------------------------
# Config key parity: DEFAULT_CONFIG <-> config.example.yaml
# ---------------------------------------------------------------------------


def test_default_config_has_baselines_key(jrc):
    assert "baselines" in jrc.DEFAULT_CONFIG["compliance"]
    assert jrc.DEFAULT_CONFIG["compliance"]["baselines"] == []


def test_config_example_compliance_keys_match_default(jrc):
    import yaml

    repo_root = Path(jrc.__file__).resolve().parent
    example = yaml.safe_load((repo_root / "config.example.yaml").read_text(encoding="utf-8"))
    default_keys = set(jrc.DEFAULT_CONFIG["compliance"].keys())
    example_keys = set(example["compliance"].keys())
    # The example must not introduce phantom keys, and must surface baselines.
    assert example_keys <= default_keys
    assert "baselines" in example_keys


# ---------------------------------------------------------------------------
# CoreDashboard sheets: render + skip-when-no-data
# ---------------------------------------------------------------------------


def _dashboard(jrc, tmp_path: Path, compliance: dict[str, Any], rows, data_dir=None):
    import xlsxwriter
    from unittest.mock import MagicMock

    cfg = _config(jrc, tmp_path, compliance)
    bridge = MagicMock()
    bridge.ea_results_report.return_value = rows
    bridge._data_dir = data_dir or (tmp_path / "jamf-cli-data")
    wb_path = str(tmp_path / "mscp.xlsx")
    wb = xlsxwriter.Workbook(wb_path, {"remove_timezone": True})
    fmts = jrc._build_formats(wb)
    dash = jrc.CoreDashboard(cfg, bridge, wb, fmts)
    return dash, wb, wb_path


def test_mscp_sheet_renders_band_row(jrc, tmp_path):
    import openpyxl

    rows = [_row(0, device="d1"), _row(5, device="d2"), _row(0, device="d3")]
    dash, wb, wb_path = _dashboard(
        jrc, tmp_path, {"failures_count_column": "Failed Count", "baseline_label": "mSCP"}, rows
    )
    dash._write_mscp_compliance()
    wb.close()
    sheet = openpyxl.load_workbook(wb_path)["mSCP Compliance"]
    text = "\n".join(
        str(c.value) for col in sheet.iter_cols() for c in col if c.value is not None
    )
    assert "mSCP" in text
    assert "Pass (0)" in text


def test_mscp_sheet_skips_when_no_baseline(jrc, tmp_path):
    dash, wb, _ = _dashboard(jrc, tmp_path, {"baselines": []}, [_row(0)])
    with pytest.raises(RuntimeError):
        dash._write_mscp_compliance()
    wb.close()


def test_mscp_sheet_skips_when_no_rows(jrc, tmp_path):
    dash, wb, _ = _dashboard(
        jrc, tmp_path, {"failures_count_column": "Failed Count"}, []
    )
    with pytest.raises(RuntimeError):
        dash._write_mscp_compliance()
    wb.close()


def test_mscp_sheet_renders_nodata_row_when_ea_unpopulated(jrc, tmp_path):
    # Device exists in the universe but has no row for this baseline's EA.
    # The sheet should render it as a No Data row (universe total > 0), with
    # Compliance % shown as "—" — mirroring Swift's mscpBandsMap keeping the
    # baseline when totalDevices > 0.
    import openpyxl

    rows = [_row(0, ea_name="Other", device="d1")]
    dash, wb, wb_path = _dashboard(
        jrc, tmp_path, {"failures_count_column": "Failed Count", "baseline_label": "mSCP"}, rows
    )
    dash._write_mscp_compliance()
    wb.close()
    sheet = openpyxl.load_workbook(wb_path)["mSCP Compliance"]
    values = [c.value for col in sheet.iter_cols() for c in col if c.value is not None]
    assert "mSCP" in values
    assert "—" in values  # compliance pct undefined (no devices with data)


def test_compliance_trend_sheet_skips_when_no_snapshots(jrc, tmp_path):
    dash, wb, _ = _dashboard(
        jrc, tmp_path, {"failures_count_column": "Failed Count"}, [],
        data_dir=tmp_path / "empty-data",
    )
    with pytest.raises(RuntimeError):
        dash._write_mscp_compliance_trend()
    wb.close()


def test_compliance_trend_sheet_renders_per_date_rows(jrc, tmp_path):
    import openpyxl

    data_dir = tmp_path / "jamf-cli-data"
    results_dir = data_dir / "ea-results"
    results_dir.mkdir(parents=True)
    (results_dir / "ea-results_2026-06-01.json").write_text(
        json.dumps([_row(0, device="d1"), _row(5, device="d2")]), encoding="utf-8"
    )
    (results_dir / "ea-results_2026-06-02.json").write_text(
        json.dumps([_row(0, device="d1")]), encoding="utf-8"
    )
    dash, wb, wb_path = _dashboard(
        jrc, tmp_path, {"failures_count_column": "Failed Count", "baseline_label": "mSCP"},
        [], data_dir=data_dir,
    )
    dash._write_mscp_compliance_trend()
    wb.close()
    sheet = openpyxl.load_workbook(wb_path)["Compliance Trend"]
    dates = [c.value for col in sheet.iter_cols(min_col=1, max_col=1) for c in col]
    assert "2026-06-01" in dates
    assert "2026-06-02" in dates


# ---------------------------------------------------------------------------
# Charts: band donut + band trend stackplot
# ---------------------------------------------------------------------------


def test_mscp_band_charts_produce_pngs(jrc, tmp_path):
    if not jrc._load_matplotlib():
        pytest.skip("matplotlib not installed")
    import xlsxwriter

    data_dir = tmp_path / "jamf-cli-data"
    results_dir = data_dir / "ea-results"
    results_dir.mkdir(parents=True)
    (results_dir / "ea-results_2026-06-01.json").write_text(
        json.dumps([_row(0, device="d1"), _row(20, device="d2")]), encoding="utf-8"
    )
    (results_dir / "ea-results_2026-06-02.json").write_text(
        json.dumps([_row(0, device="d1"), _row(0, device="d2")]), encoding="utf-8"
    )
    cfg = _config(
        jrc, tmp_path, {"failures_count_column": "Failed Count", "baseline_label": "NIST"}
    )
    out_dir = tmp_path / "out"
    out_dir.mkdir()
    wb = xlsxwriter.Workbook(str(out_dir / "charts.xlsx"), {"remove_timezone": True})
    gen = jrc.ChartGenerator(
        config=cfg,
        csv_path=None,
        historical_dir=None,
        output_dir=out_dir,
        workbook=wb,
        jamf_cli_dir=data_dir,
        output_stem="report",
        embed_in_workbook=False,
    )
    gen._chart_dir = out_dir
    paths = gen._generate_mscp_band_charts()
    wb.close()
    names = [Path(p).name for p in paths]
    assert any("mscp_band_donut" in n for n in names)
    assert any("mscp_band_trend" in n for n in names)
    for p in paths:
        assert Path(p).is_file()


def test_mscp_band_charts_empty_when_no_baseline(jrc, tmp_path):
    if not jrc._load_matplotlib():
        pytest.skip("matplotlib not installed")
    import xlsxwriter

    cfg = _config(jrc, tmp_path, {"baselines": []})
    out_dir = tmp_path / "out"
    out_dir.mkdir()
    wb = xlsxwriter.Workbook(str(out_dir / "charts.xlsx"), {"remove_timezone": True})
    gen = jrc.ChartGenerator(
        config=cfg,
        csv_path=None,
        historical_dir=None,
        output_dir=out_dir,
        workbook=wb,
        jamf_cli_dir=tmp_path / "jamf-cli-data",
        output_stem="report",
        embed_in_workbook=False,
    )
    gen._chart_dir = out_dir
    assert gen._generate_mscp_band_charts() == []
    wb.close()
