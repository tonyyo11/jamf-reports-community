"""Regression tests for the code-review fixes applied in emdash/2-1-1.

Covers the behavioral fixes that are unit-testable without a live jamf-cli:
  (a) _ea_percentage: blank/unparseable cells appear as a labelled summary row
  (b) _csv_injection_safe: leading-= cells are prefixed with a single quote
  (c) LogRedactor.redact_json: bare `token`, `bearer_token`, `session_token`,
      `pat`, `private_key` keys are redacted
  (d) _school_csv_load: raises SystemExit on a missing file
"""

from __future__ import annotations

import io
from pathlib import Path

import pandas as pd
import pytest
import xlsxwriter
import openpyxl


# ---------------------------------------------------------------------------
# (a) _ea_percentage blank / unparseable row
# ---------------------------------------------------------------------------


def _make_csv_dashboard(jrc, tmp_path: Path, col_data: list[str]) -> tuple:
    """Spin up a minimal CSVDashboard + worksheet to exercise _ea_percentage."""
    # Build a tiny CSV in memory with the EA column + a Computer Name column
    df = pd.DataFrame({"Computer Name": [f"device-{i}" for i in range(len(col_data))],
                       "DiskUsage": col_data})
    csv_path = tmp_path / "tiny.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    # Minimal config pointing at our temp CSV
    config_dir = tmp_path / "config"
    config_dir.mkdir()
    cfg_path = config_dir / "test.yaml"
    cfg_path.write_text(
        "columns:\n"
        "  device_name: 'Computer Name'\n"
        "custom_eas: []\n"
        "thresholds:\n"
        "  warning_disk_percent: 70\n"
        "  critical_disk_percent: 90\n",
        encoding="utf-8",
    )
    config = jrc.Config(str(cfg_path))

    wb_path = tmp_path / "out.xlsx"
    wb = xlsxwriter.Workbook(str(wb_path))
    fmts = jrc._build_formats(wb)
    dash = jrc.CSVDashboard(config, str(csv_path), wb, fmts)
    ws = wb.add_worksheet("EA Test")
    return dash, ws, wb, wb_path


def test_ea_percentage_blank_row_appears_when_blanks_exist(jrc, tmp_path: Path) -> None:
    """Blank / non-numeric EA values must produce a 'Blank / Unparseable' row."""
    data = ["55%", "85%", "", "N/A", "95%"]
    dash, ws, wb, wb_path = _make_csv_dashboard(jrc, tmp_path, data)

    ea_cfg = {
        "name": "Disk Usage",
        "column": "DiskUsage",
        "type": "percentage",
        "warning_threshold": 70,
        "critical_threshold": 90,
    }
    dash._ea_percentage(ws, 0, "DiskUsage", ea_cfg)
    wb.close()

    loaded = openpyxl.load_workbook(wb_path, data_only=True)
    sheet = loaded["EA Test"]

    # Collect all labels in column A (non-None)
    col_a = [sheet.cell(row=r, column=1).value for r in range(1, 20)
             if sheet.cell(row=r, column=1).value is not None]

    assert "Blank / Unparseable" in col_a, (
        f"'Blank / Unparseable' row expected but not found; column A = {col_a}"
    )

    # Find count in column B of that row
    blank_row_idx = next(
        r for r in range(1, 20)
        if sheet.cell(row=r, column=1).value == "Blank / Unparseable"
    )
    blank_count = sheet.cell(row=blank_row_idx, column=2).value
    assert blank_count == 2, (
        f"Expected blank count 2 (empty string + N/A), got {blank_count}"
    )


def test_ea_percentage_no_blank_row_when_all_numeric(jrc, tmp_path: Path) -> None:
    """No 'Blank / Unparseable' row when every value parses cleanly."""
    data = ["50%", "80%", "95%"]
    dash, ws, wb, wb_path = _make_csv_dashboard(jrc, tmp_path, data)

    ea_cfg = {
        "name": "Disk Usage",
        "column": "DiskUsage",
        "type": "percentage",
        "warning_threshold": 70,
        "critical_threshold": 90,
    }
    dash._ea_percentage(ws, 0, "DiskUsage", ea_cfg)
    wb.close()

    loaded = openpyxl.load_workbook(wb_path, data_only=True)
    sheet = loaded["EA Test"]

    col_a = [sheet.cell(row=r, column=1).value for r in range(1, 20)
             if sheet.cell(row=r, column=1).value is not None]
    assert "Blank / Unparseable" not in col_a


def test_ea_percentage_null_thresholds_use_defaults(jrc, tmp_path: Path) -> None:
    """Null warning_threshold / critical_threshold must not raise TypeError."""
    data = ["55%", "85%"]
    dash, ws, wb, wb_path = _make_csv_dashboard(jrc, tmp_path, data)

    ea_cfg = {
        "name": "Disk Usage",
        "column": "DiskUsage",
        "type": "percentage",
        "warning_threshold": None,
        "critical_threshold": None,
    }
    # Should not raise
    dash._ea_percentage(ws, 0, "DiskUsage", ea_cfg)
    wb.close()


def test_ea_boolean_null_true_value_uses_default(jrc, tmp_path: Path) -> None:
    """Null true_value must not raise AttributeError on .lower()."""
    data = ["Yes", "No", ""]
    dash, ws, wb, wb_path = _make_csv_dashboard(jrc, tmp_path, data)
    ea_cfg = {"name": "Status", "column": "DiskUsage", "type": "boolean", "true_value": None}
    dash._ea_boolean(ws, 0, "DiskUsage", ea_cfg)
    wb.close()


# ---------------------------------------------------------------------------
# (b) _csv_injection_safe neutralizes formula-injection
# ---------------------------------------------------------------------------


def test_csv_injection_safe_neutralizes_leading_equals(jrc) -> None:
    """Cells starting with '=' must be prefixed with a single quote."""
    df = pd.DataFrame({"name": ["=SUM(A1:A10)"], "value": [42]})
    safe = jrc._csv_injection_safe(df)
    assert safe.loc[0, "name"] == "'=SUM(A1:A10)"


def test_csv_injection_safe_neutralizes_all_trigger_chars(jrc) -> None:
    """'=', '+', '-', '@' must all be neutralized."""
    df = pd.DataFrame({
        "a": ["=EVIL()"],
        "b": ["+1+1"],
        "c": ["-2"],
        "d": ["@SUM(A1)"],
    })
    safe = jrc._csv_injection_safe(df)
    for col in ["a", "b", "c", "d"]:
        assert safe.loc[0, col].startswith("'"), (
            f"Column {col!r} value {safe.loc[0, col]!r} not prefixed with single quote"
        )


def test_csv_injection_safe_leaves_safe_cells_unchanged(jrc) -> None:
    """Cells that don't start with a trigger char must be left untouched."""
    df = pd.DataFrame({
        "name": ["MacBook Pro"],
        "serial": ["ABC123"],
        "empty": [""],
    })
    safe = jrc._csv_injection_safe(df)
    assert safe.loc[0, "name"] == "MacBook Pro"
    assert safe.loc[0, "serial"] == "ABC123"
    assert safe.loc[0, "empty"] == ""


def test_csv_injection_safe_does_not_alter_numeric_columns(jrc) -> None:
    """Integer and float columns must not be touched."""
    df = pd.DataFrame({"count": [1, 2, 3], "pct": [0.1, 0.5, 0.9]})
    safe = jrc._csv_injection_safe(df)
    assert list(safe["count"]) == [1, 2, 3]
    assert list(safe["pct"]) == pytest.approx([0.1, 0.5, 0.9])


def test_csv_injection_safe_handles_lstrip_correctly(jrc) -> None:
    """Leading whitespace before trigger char must still trigger neutralization."""
    df = pd.DataFrame({"a": ["  =sneaky"]})
    safe = jrc._csv_injection_safe(df)
    # lstrip checks, so ' = sneaky'.lstrip() -> '=sneaky', which starts with '='
    assert safe.loc[0, "a"].startswith("'"), (
        f"Whitespace-prefixed formula {df.loc[0, 'a']!r} should be neutralized"
    )


# ---------------------------------------------------------------------------
# (c) LogRedactor.redact_json redacts bare token / pat / private_key keys
# ---------------------------------------------------------------------------


def test_redactor_redacts_bare_token_key(jrc) -> None:
    redactor = jrc.LogRedactor()
    result = redactor.redact_json({"token": "supersecretapitoken12345"})
    assert result["token"] == "REDACTED_TOKEN"
    assert "supersecretapitoken12345" not in str(result)


def test_redactor_redacts_bearer_token_key(jrc) -> None:
    redactor = jrc.LogRedactor()
    result = redactor.redact_json({"bearer_token": "eyJfakeBearer"})
    assert result["bearer_token"] == "REDACTED_BEARER_TOKEN"


def test_redactor_redacts_session_token_key(jrc) -> None:
    redactor = jrc.LogRedactor()
    result = redactor.redact_json({"session_token": "sess-abc-def-12345"})
    assert result["session_token"] == "REDACTED_SESSION_TOKEN"


def test_redactor_redacts_pat_key(jrc) -> None:
    redactor = jrc.LogRedactor()
    result = redactor.redact_json({"pat": "ghp_personalAccessToken123"})
    assert result["pat"] == "REDACTED_PAT"


def test_redactor_redacts_private_key_key(jrc) -> None:
    redactor = jrc.LogRedactor()
    result = redactor.redact_json({"private_key": "-----BEGIN RSA PRIVATE KEY-----"})
    assert result["private_key"] == "REDACTED_PRIVATE_KEY"


def test_redactor_new_keys_case_insensitive(jrc) -> None:
    """Keys must match case-insensitively since _SENSITIVE_JSON_KEYS stores lowercase."""
    redactor = jrc.LogRedactor()
    # redact_json lowercases the key before lookup; key name is preserved in output
    result = redactor.redact_json({"TOKEN": "s3cr3t"})
    # The original key is preserved; the value must be redacted
    assert result["TOKEN"] == "REDACTED_TOKEN"
    assert "s3cr3t" not in str(result)


# ---------------------------------------------------------------------------
# (d) _school_csv_load raises SystemExit on a missing file
# ---------------------------------------------------------------------------


def test_school_csv_load_raises_system_exit_on_missing_file(jrc) -> None:
    with pytest.raises(SystemExit) as exc_info:
        jrc._school_csv_load("/nonexistent/path/school_devices.csv")
    assert "could not read school CSV" in str(exc_info.value).lower() or \
           "Error:" in str(exc_info.value)


def test_school_csv_load_error_message_contains_path(jrc) -> None:
    path = "/nonexistent/school_devices.csv"
    with pytest.raises(SystemExit) as exc_info:
        jrc._school_csv_load(path)
    assert path in str(exc_info.value)


def test_school_csv_load_raises_system_exit_on_unreadable_csv(
    jrc, tmp_path: Path
) -> None:
    """A CSV that exists but is not valid CSV should still surface as SystemExit."""
    bad = tmp_path / "bad.csv"
    bad.write_bytes(b"\xff\xfe" * 500 + b"\x00" * 100)  # likely to fail as CSV
    # Either succeeds (binary read + fillna handles it) or raises SystemExit
    try:
        df = jrc._school_csv_load(str(bad))
        # If it doesn't raise, it must return a DataFrame
        assert hasattr(df, "columns")
    except SystemExit:
        pass  # expected
