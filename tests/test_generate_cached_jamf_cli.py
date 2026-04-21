"""Integration tests for cached jamf-cli workbook generation."""

from __future__ import annotations

import openpyxl
import pytest


@pytest.mark.integration
@pytest.mark.filterwarnings("ignore:No artists with labels found to put in legend")
def test_generate_from_committed_cached_jamf_cli_data(
    monkeypatch,
    config_factory,
    tmp_path,
    jrc,
) -> None:
    config = config_factory("dummy.yaml")
    config._data["jamf_cli"]["enabled"] = True
    config._data["jamf_cli"]["allow_live_overview"] = False
    monkeypatch.setattr(jrc, "_find_jamf_cli_binary", lambda: None)

    report_path = jrc.cmd_generate(config, None, str(tmp_path / "cached-jamf-cli.xlsx"))

    assert report_path.exists()
    workbook = openpyxl.load_workbook(report_path, data_only=False)
    expected = {
        "Fleet Overview",
        "Mobile Fleet Summary",
        "Security Posture",
        "Inventory Summary",
        "Device Compliance",
        "Package Lifecycle",
        "Active Devices",
        "Patch Compliance",
        "Smart Groups",
        "Update Status",
        "Update Failures",
        "Report Sources",
    }
    assert expected.issubset(set(workbook.sheetnames))

    active_sheet = workbook["Active Devices"]
    col_a = [active_sheet[f"A{r}"].value for r in range(1, 10)]
    assert "Active Devices" in col_a
    assert "Active Window (days)" in col_a
    assert "Total Devices" in col_a
    assert "Active Devices" in col_a
    assert "Inactive Devices" in col_a
    assert "Active Ratio %" in col_a

    patch_sheet = workbook["Patch Compliance"]
    patch_headers = [patch_sheet.cell(row=4, column=c).value for c in range(1, 11)]
    assert "Adjusted Up To Date" in patch_headers or "Adjusted Installed" in patch_headers

    sheet = workbook["Smart Groups"]
    assert sheet["A4"].value == "Group Name"
    assert sheet["A5"].value == "3PL AMR Security Profile - 2.0 60 Minute Screensaver"
    assert sheet["B5"].value == "Computer"
    assert sheet["C5"].value == "No"
    assert sheet["D5"].value == 0
    assert sheet["G5"].value == "Zero members"

    package_sheet = workbook["Package Lifecycle"]
    assert package_sheet["A1"].value == "Package Lifecycle"
    assert "jamf-cli pro packages list" in str(package_sheet["A2"].value)
    assert package_sheet["A4"].value == "Total Packages"
    assert package_sheet["B4"].value == 12
    assert package_sheet["A6"].value == "Package Name"
    assert package_sheet["A7"].value == "AdobeAcrobatPro11CC_2014-08-06.pkg"
    notes = [package_sheet[f"G{row}"].value for row in range(7, 19)]
    assert "Reboot required" in notes


@pytest.mark.integration
def test_generate_self_contained_html_from_cached_jamf_cli_data(
    monkeypatch,
    config_factory,
    tmp_path,
    jrc,
) -> None:
    config = config_factory("dummy.yaml")
    config._data["jamf_cli"]["enabled"] = True
    config._data["jamf_cli"]["allow_live_overview"] = False
    monkeypatch.setattr(jrc, "_find_jamf_cli_binary", lambda: None)

    html_path = tmp_path / "cached-jamf-cli.html"
    jrc.cmd_html(config, str(html_path), no_open=True)

    html = html_path.read_text(encoding="utf-8")
    assert "cdn.jsdelivr.net" not in html
    assert '<script src="' not in html
    assert "Report Sources" in html
    assert "const safeCsvValue" in html
    assert "innerHTML =" not in html
    assert '<svg class="trend-svg"' in html


@pytest.mark.integration
def test_generate_cached_jamf_cli_respects_sheets_skip(
    monkeypatch,
    config_factory,
    tmp_path,
    jrc,
) -> None:
    config = config_factory("dummy.yaml")
    config._data["jamf_cli"]["enabled"] = True
    config._data["jamf_cli"]["allow_live_overview"] = False
    config._data["sheets"] = {
        "skip": ["Update Status", "Report Sources"],
    }
    monkeypatch.setattr(jrc, "_find_jamf_cli_binary", lambda: None)

    report_path = jrc.cmd_generate(config, None, str(tmp_path / "cached-jamf-cli-skip.xlsx"))

    workbook = openpyxl.load_workbook(report_path, data_only=False)
    assert "Update Status" not in workbook.sheetnames
    assert "Report Sources" not in workbook.sheetnames
    assert "Fleet Overview" in workbook.sheetnames


@pytest.mark.integration
def test_generate_warns_for_unknown_sheet_skip(
    monkeypatch,
    config_factory,
    tmp_path,
    capsys,
    jrc,
) -> None:
    config = config_factory("dummy.yaml")
    config._data["jamf_cli"]["enabled"] = True
    config._data["jamf_cli"]["allow_live_overview"] = False
    config._data["sheets"] = {
        "skip": ["Not A Real Sheet"],
    }
    monkeypatch.setattr(jrc, "_find_jamf_cli_binary", lambda: None)

    jrc.cmd_generate(config, None, str(tmp_path / "cached-jamf-cli-warning.xlsx"))

    captured = capsys.readouterr()
    assert "sheets.skip: unknown sheet 'Not A Real Sheet'" in captured.out


@pytest.mark.integration
def test_html_cleanup_section_with_cached_detail(
    monkeypatch,
    config_factory,
    tmp_path,
    jrc,
) -> None:
    """Cleanup Analysis section renders when per-policy/profile detail is cached."""
    config = config_factory("dummy.yaml")
    config._data["jamf_cli"]["enabled"] = True
    config._data["jamf_cli"]["allow_live_overview"] = False
    monkeypatch.setattr(jrc, "_find_jamf_cli_binary", lambda: None)

    html_path = tmp_path / "cleanup-test.html"
    jrc.cmd_html(config, str(html_path), no_open=True)

    html = html_path.read_text(encoding="utf-8")
    # Section header appears when detail cache exists
    assert "Cleanup Analysis" in html
    # Disabled policy from fixture (policy id=2, enabled=false)
    assert "Adobe InCopy CC 2014" in html
    # DiffMerge is enabled + has no scope targets — should appear as unscoped
    assert "DiffMerge" in html
    # WiFi Test profile has no scope targets — should appear as unscoped
    assert "WiFi Test" in html
    # AdobeInCopyCC2014 pkg (id=2) is referenced by policy 2, DiffMerge420 (id=1)
    # is referenced by policy 14 — remaining packages should be listed as unused
    # (at minimum the section renders without error)
    assert "Unused Packages" in html
    assert "Unused Scripts" in html


@pytest.mark.integration
def test_html_cleanup_section_absent_without_detail_cache(
    monkeypatch,
    config_factory,
    tmp_path,
    fixtures_root,
    jrc,
) -> None:
    """Cleanup Analysis section is omitted when no per-policy detail is cached."""
    import shutil

    config = config_factory("dummy.yaml")
    config._data["jamf_cli"]["enabled"] = True
    config._data["jamf_cli"]["allow_live_overview"] = False
    # Point data_dir to a temp dir that has list data but no per-ID subdirs
    stripped_dir = tmp_path / "stripped-cache"
    stripped_dir.mkdir()
    src = fixtures_root / "jamf-cli-data"
    for subdir in src.iterdir():
        if subdir.is_dir() and subdir.name not in ("classic-policies", "classic-macos-profiles"):
            shutil.copytree(subdir, stripped_dir / subdir.name)
        elif subdir.is_dir():
            # Copy only the list-level files, not per-ID subdirectories
            dest = stripped_dir / subdir.name
            dest.mkdir()
            for f in subdir.iterdir():
                if f.is_file():
                    shutil.copy2(f, dest / f.name)

    config._data["jamf_cli"]["data_dir"] = str(stripped_dir)
    monkeypatch.setattr(jrc, "_find_jamf_cli_binary", lambda: None)

    html_path = tmp_path / "no-cleanup.html"
    jrc.cmd_html(config, str(html_path), no_open=True)

    html = html_path.read_text(encoding="utf-8")
    # Section should be absent when no detail cache is present
    assert "Cleanup Analysis" not in html
